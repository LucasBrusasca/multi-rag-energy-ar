"""Two readers, one grid: Docling for PDF, openpyxl for spreadsheets.

Everything downstream works on `SegmentoTabla`, so the header, unit,
continuation and fact logic is written once and is identical for both. What
changes is only how the grid is read.

The spreadsheet is NOT read with Docling. Docling turned the audited workbook
into 251 fragments, 167 of them single-column, and the income statement ended up
split across three of them (`DIAGNOSTICO.md`, section 2.4). openpyxl reads it
intact, with no model involved.

[ES] Dos lectores, una grilla: Docling para PDF, openpyxl para planillas.

Todo lo que sigue trabaja sobre `SegmentoTabla`, asi que la logica de
encabezado, unidad, continuacion y hechos se escribe una sola vez y es identica
para los dos. Lo unico que cambia es como se lee la grilla.

La planilla NO se lee con Docling. Docling convirtio el libro auditado en 251
fragmentos, 167 de una sola columna, y el estado de resultados quedo repartido
en tres de ellos (`DIAGNOSTICO.md`, seccion 2.4). openpyxl lo lee integro, sin
ningun modelo de por medio.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterator, Optional

from multirag.ingestion.tablas.continuidad import enlazar_continuaciones
from multirag.ingestion.tablas.grilla import (
    NUMERO,
    TEXTO,
    VACIO,
    clasificar,
    contrastar_con_parser,
    expandir,
    inferir_banda_encabezado,
    inferir_columna_etiqueta,
    normalizar,
)
from multirag.ingestion.tablas.modelo import (
    Celda,
    SegmentoTabla,
    uid_segmento,
    uid_tabla,
)
from multirag.ingestion.tablas.semantica import detectar_unidad, huele_a_unidad


# How far back in reading order a unit declaration may be picked up. The scan
# always stops at the previous table, so a unit is never stolen from another.
# [ES] Hasta donde hacia atras en el orden de lectura se puede tomar una
# declaracion de unidad. El barrido siempre frena en la tabla anterior, asi que
# nunca se le roba la unidad a otra.
VENTANA_TEXTO_ADYACENTE = 6


def _identidad(identidad: Optional[dict]) -> dict:
    identidad = identidad or {}
    return {
        "document_id": identidad.get("document_id"),
        "artifact_id": identidad.get("artifact_id"),
        "fuente": identidad.get("fuente"),
        "entidad": identidad.get("entidad") or identidad.get("emisor") or identidad.get("fuente"),
    }


def preparar_segmento(
    segmento: SegmentoTabla, candidatos_unidad: list | None = None
) -> SegmentoTabla:
    """Run the shared inference over an already-populated grid.

    Public so it can be exercised on a hand-built grid, with no parser in the
    way: the header, unit and continuation rules are the part that has to be
    provable.

    [ES] Corre la inferencia compartida sobre una grilla ya poblada.

    Publica para poder ejercitarla sobre una grilla armada a mano, sin ningun
    parser en el medio: las reglas de encabezado, unidad y continuacion son la
    parte que hay que poder demostrar.
    """
    candidatos_unidad = list(candidatos_unidad or [])
    grilla = expandir(segmento)
    segmento.columna_etiqueta = inferir_columna_etiqueta(
        grilla, segmento.num_rows, segmento.num_cols
    )

    # A table with no amounts is out of scope here. Its header cannot be told
    # from its data by shape, and it has no figure to make interpretable: it
    # stays with the text chunk, which already serves it well.
    # [ES] Una tabla sin importes queda fuera de alcance. Su encabezado no se
    # distingue de sus datos por la forma, y no tiene ninguna cifra que volver
    # interpretable: se queda con el chunk de texto, que ya la sirve bien.
    tiene_numeros = any(clasificar(c.texto) == NUMERO for c in segmento.celdas)
    if not tiene_numeros:
        segmento.banda_encabezado = ()
        segmento.avisar("tabla_sin_valores_numericos")
        segmento.anotar("fuera_de_alcance:sin_importes")
        return segmento

    segmento.banda_encabezado = inferir_banda_encabezado(
        grilla, segmento.num_rows, segmento.num_cols, segmento.columna_etiqueta
    )
    contrastar_con_parser(segmento, grilla)

    # A header-band cell that declares a unit outranks any nearby paragraph.
    # [ES] Una celda de la banda de encabezado que declara unidad pesa mas que
    # cualquier parrafo cercano.
    del_encabezado = [
        (normalizar(grilla[(f, c)].texto), "celda_encabezado", f"r{f}c{c}")
        for f in segmento.banda_encabezado
        for c in range(segmento.num_cols)
        if (f, c) in grilla and huele_a_unidad(grilla[(f, c)].texto)
    ]
    if segmento.banda_encabezado:
        titulo = normalizar(
            grilla[(segmento.banda_encabezado[0], segmento.columna_etiqueta)].texto
            if (segmento.banda_encabezado[0], segmento.columna_etiqueta) in grilla
            else ""
        )
        segmento.titulo_inferido = titulo or None

    segmento.unidad, avisos = detectar_unidad(del_encabezado + candidatos_unidad)
    for aviso in avisos:
        segmento.avisar(aviso)
    if not segmento.banda_encabezado:
        segmento.avisar("sin_encabezado_propio")
    return segmento


# --------------------------------------------------------------------------
# Docling (PDF)
# --------------------------------------------------------------------------

def _orden_de_lectura(documento: dict) -> Iterator[str]:
    """Flatten `body` into reading order, descending into groups.

    [ES] Aplana `body` a orden de lectura, entrando en los grupos.
    """
    grupos = {g["self_ref"]: g for g in documento.get("groups", [])}

    def recorrer(nodo: dict, visto: set[str]) -> Iterator[str]:
        for hijo in nodo.get("children", []) or []:
            ref = hijo.get("$ref") if isinstance(hijo, dict) else hijo
            if not ref or ref in visto:
                continue
            visto.add(ref)
            if ref in grupos:
                yield from recorrer(grupos[ref], visto)
            else:
                yield ref

    yield from recorrer(documento.get("body", {}), set())


def segmentos_desde_docling(
    documento: dict,
    parser_version: str,
    identidad: Optional[dict] = None,
) -> list[SegmentoTabla]:
    """Build one segment per Docling table, in reading order.

    `documento` is `DoclingDocument.export_to_dict()`: taking the dict instead
    of the live object is what lets a conversion be cached and re-analysed
    without reconverting, which is how the audit already works.

    [ES] Construye un segmento por tabla de Docling, en orden de lectura.

    `documento` es `DoclingDocument.export_to_dict()`: tomar el dict en lugar
    del objeto vivo es lo que permite cachear una conversion y re-analizarla sin
    reconvertir, que es como ya funciona la auditoria.
    """
    ident = _identidad(identidad)
    textos = {t["self_ref"]: t for t in documento.get("texts", [])}
    tablas = {t["self_ref"]: t for t in documento.get("tables", [])}
    orden = list(_orden_de_lectura(documento))

    segmentos: list[SegmentoTabla] = []
    for posicion, ref in enumerate(orden):
        if ref not in tablas:
            continue
        crudo = tablas[ref]
        datos = crudo.get("data") or {}
        prov = crudo.get("prov") or []
        paginas = tuple(
            sorted({p["page_no"] for p in prov if p.get("page_no") is not None})
        )

        celdas = [
            Celda(
                fila=c["start_row_offset_idx"],
                col=c["start_col_offset_idx"],
                texto=c.get("text", ""),
                fila_span=max(1, c.get("row_span", 1)),
                col_span=max(1, c.get("col_span", 1)),
                es_encabezado_col_parser=bool(c.get("column_header")),
                es_encabezado_fila_parser=bool(c.get("row_header")),
                es_seccion_parser=bool(c.get("row_section")),
                bbox=c.get("bbox"),
                pagina=paginas[0] if paginas else None,
            )
            for c in datos.get("table_cells", [])
        ]

        caption = " ".join(
            normalizar(textos[r["$ref"]]["text"])
            for r in (crudo.get("captions") or [])
            if isinstance(r, dict) and r.get("$ref") in textos
        ).strip()

        segmento = SegmentoTabla(
            table_segment_uid=uid_segmento(ident["artifact_id"] or "", ref),
            table_uid=uid_tabla(ident["artifact_id"] or "", ref),
            continuation_of=None,
            parser="docling",
            parser_version=parser_version,
            parser_marca_encabezados=True,
            ancla=ref,
            source_pages=paginas,
            caption=caption or None,
            num_rows=datos.get("num_rows", 0),
            num_cols=datos.get("num_cols", 0),
            celdas=celdas,
            **ident,
        )

        candidatos = []
        if caption:
            candidatos.append((caption, "caption", ref))
        for texto_crudo, referencia in _textos_previos(orden, posicion, textos, tablas):
            if huele_a_unidad(texto_crudo):
                candidatos.append((texto_crudo, "texto_adyacente", referencia))
                break

        segmentos.append(preparar_segmento(segmento, candidatos))

    enlazar_continuaciones(segmentos)
    return segmentos


def _textos_previos(
    orden: list[str], posicion: int, textos: dict, tablas: dict
) -> Iterator[tuple[str, str]]:
    """Walk backwards from a table, stopping at the previous table.

    Stopping there is the whole point: it is what keeps the unit line of one
    table from being attributed to the next.

    [ES] Camina hacia atras desde una tabla y frena en la tabla anterior.

    Frenar ahi es todo el punto: es lo que evita que la linea de unidad de una
    tabla se le atribuya a la siguiente.
    """
    vistos = 0
    for indice in range(posicion - 1, -1, -1):
        ref = orden[indice]
        if ref in tablas:
            return
        if ref not in textos:
            continue
        vistos += 1
        if vistos > VENTANA_TEXTO_ADYACENTE:
            return
        yield normalizar(textos[ref]["text"]), ref


# --------------------------------------------------------------------------
# openpyxl (spreadsheets)
# --------------------------------------------------------------------------

def _valores_de_hoja(hoja) -> tuple[dict[tuple[int, int], Any], dict[tuple[int, int], tuple[int, int]]]:
    """Sheet values with merged ranges projected onto every covered cell.

    A merged range holds its value only in the top-left cell; without this
    projection, `First quarter` spanning D2:F2 would be the header of D and of
    nothing else.

    [ES] Valores de la hoja con los rangos combinados proyectados sobre todas
    las celdas que cubren.

    Un rango combinado guarda su valor solo en la celda superior izquierda; sin
    esta proyeccion, `First quarter` sobre D2:F2 seria encabezado de D y de nada
    mas.
    """
    valores: dict[tuple[int, int], Any] = {}
    for fila in hoja.iter_rows():
        for celda in fila:
            if celda.value is not None:
                valores[(celda.row, celda.column)] = celda.value

    spans: dict[tuple[int, int], tuple[int, int]] = {}
    for rango in hoja.merged_cells.ranges:
        origen = (rango.min_row, rango.min_col)
        spans[origen] = (
            rango.max_row - rango.min_row + 1,
            rango.max_col - rango.min_col + 1,
        )
        valor = valores.get(origen)
        if valor is None:
            continue
        for f in range(rango.min_row, rango.max_row + 1):
            for c in range(rango.min_col, rango.max_col + 1):
                valores.setdefault((f, c), valor)
    return valores, spans


def _clase_de_fila_excel(
    valores: dict, fila: int, columnas: list[int], columna_etiqueta: int
) -> str:
    clases = [
        clasificar(valores.get((fila, c)))
        for c in columnas
        if c != columna_etiqueta and (fila, c) in valores
    ]
    utiles = [k for k in clases if k != VACIO]
    tiene_etiqueta = bool(normalizar(valores.get((fila, columna_etiqueta))))
    if not utiles:
        return "SECCION" if tiene_etiqueta else "VACIA"
    return "ENCABEZADO" if NUMERO not in utiles else "DATO"


def segmentos_desde_excel(
    ruta: Path,
    parser_version: str,
    identidad: Optional[dict] = None,
    hojas: Optional[tuple[str, ...]] = None,
) -> list[SegmentoTabla]:
    """One segment per header block of each sheet.

    A sheet is not one table: the audited workbook restates its header five
    times in a single sheet, once per reconciliation. Each restatement opens a
    new block, and they are NOT continuations of one another, because they
    describe different concepts.

    [ES] Un segmento por bloque de encabezado de cada hoja.

    Una hoja no es una tabla: el libro auditado repite su encabezado cinco veces
    en una sola hoja, una por conciliacion. Cada repeticion abre un bloque, y NO
    son continuacion una de otra, porque describen conceptos distintos.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    ident = _identidad(identidad)
    libro = openpyxl.load_workbook(ruta, data_only=True, read_only=False)
    segmentos: list[SegmentoTabla] = []

    for nombre in libro.sheetnames:
        if hojas and nombre not in hojas:
            continue
        hoja = libro[nombre]
        valores, spans = _valores_de_hoja(hoja)
        if not valores:
            continue

        filas = sorted({f for f, _ in valores})
        columnas = sorted({c for _, c in valores})

        etiqueta = next(
            (
                c
                for c in columnas
                if any(
                    clasificar(valores.get((f, c))) == TEXTO
                    for f in filas
                )
                and not any(clasificar(valores.get((f, c))) == NUMERO for f in filas)
            ),
            columnas[0],
        )

        bloques: list[list[int]] = []
        actual: list[int] = []
        vio_datos = False
        for fila in filas:
            clase = _clase_de_fila_excel(valores, fila, columnas, etiqueta)
            if clase == "ENCABEZADO" and (vio_datos or not actual):
                if actual:
                    bloques.append(actual)
                actual, vio_datos = [fila], False
                continue
            if not actual:
                continue
            actual.append(fila)
            if clase == "DATO":
                vio_datos = True
        if actual:
            bloques.append(actual)

        cubiertas = {f for bloque in bloques for f in bloque}
        sueltas = [
            f
            for f in filas
            if f not in cubiertas
            and any((f, c) in valores for c in columnas)
        ]

        for bloque in bloques:
            primera, ultima = bloque[0], bloque[-1]
            ancla = (
                f"{nombre!r}!{get_column_letter(columnas[0])}{primera}"
                f":{get_column_letter(columnas[-1])}{ultima}"
            )
            celdas = []
            for f in bloque:
                for c in columnas:
                    if (f, c) not in valores:
                        continue
                    fila_span, col_span = spans.get((f, c), (1, 1))
                    celdas.append(
                        Celda(
                            fila=f - primera,
                            col=c - columnas[0],
                            texto=normalizar(valores[(f, c)]),
                            fila_span=fila_span,
                            col_span=col_span,
                            valor_nativo=(
                                valores[(f, c)]
                                if isinstance(valores[(f, c)], (int, float))
                                and not isinstance(valores[(f, c)], bool)
                                else None
                            ),
                            coordenada=f"{get_column_letter(c)}{f}",
                        )
                    )

            segmento = SegmentoTabla(
                table_segment_uid=uid_segmento(ident["artifact_id"] or "", ancla),
                table_uid=uid_tabla(ident["artifact_id"] or "", ancla),
                continuation_of=None,
                parser="openpyxl",
                parser_version=parser_version,
                parser_marca_encabezados=False,
                ancla=ancla,
                source_pages=(),
                hoja=nombre,
                caption=None,
                num_rows=ultima - primera + 1,
                num_cols=columnas[-1] - columnas[0] + 1,
                celdas=celdas,
                **ident,
            )
            completo = preparar_segmento(segmento)
            if sueltas:
                # A row with data outside every block would be a silent loss.
                # [ES] Una fila con datos fuera de todo bloque seria una
                # perdida silenciosa.
                completo.avisar(f"filas_fuera_de_bloque:{sueltas[:10]}")
            segmentos.append(completo)

    libro.close()
    return segmentos
