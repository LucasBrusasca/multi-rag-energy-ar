"""Build a self-contained offline HTML interface for the human catalogue review.

WHY. The v2 spreadsheet is correct and hard to work in: 22 columns, wrapped
cells and a mental model held in the reviewer's head. Twenty-four judgements
made in a cramped grid are twenty-four chances to fill the wrong cell. This puts
one document on screen at a time, with its evidence, and asks one question.

WHAT IT IS. A single HTML file. Double-click, no server, no dependencies, no
network. The data is embedded; every style and script is inline. It reads from
the approved v2 CSV and changes nothing: not the catalogue, not the spreadsheet,
not PostgreSQL.

TEXT ONLY, NEVER innerHTML. Titles, excerpts and evidence come from real
documents, and a document is untrusted input. Everything is rendered with
`textContent` and elements built by `createElement`. Links are only created when
the scheme is http or https, so a `javascript:` URL in a catalogue field cannot
become a click target.

NOTHING IS APPLIED. The interface produces a CSV and a JSON for review. It never
writes to the catalogue, and the export is a proposal, not a change.

ANCHORING IS TREATED AS A HAZARD. The automatic classification is not shown at
first and needs a deliberate click behind a warning. Showing the classifier's
answer next to the question is how a human reference quietly becomes a copy of
the machine.

[ES] Arma una interfaz HTML autocontenida y offline para la revision humana del
catalogo.

POR QUE. La planilla v2 es correcta y es incomoda: 22 columnas, celdas con ajuste
de texto y un modelo mental sostenido en la cabeza del revisor. Veinticuatro
juicios hechos en una grilla apretada son veinticuatro oportunidades de completar
la celda equivocada. Esto pone un documento por pantalla, con su evidencia, y
hace una sola pregunta.

QUE ES. Un unico archivo HTML. Doble clic, sin servidor, sin dependencias, sin
red. Los datos van embebidos; todo estilo y script va en linea. Lee del CSV v2
aprobado y no cambia nada: ni el catalogo, ni la planilla, ni PostgreSQL.

SOLO TEXTO, NUNCA innerHTML. Titulos, extractos y evidencia vienen de documentos
reales, y un documento es entrada no confiable. Todo se renderiza con
`textContent` y elementos creados con `createElement`. Los enlaces se crean solo
si el esquema es http o https, asi que una URL `javascript:` en un campo del
catalogo no puede convertirse en algo clickeable.

NO SE APLICA NADA. La interfaz produce un CSV y un JSON para revisar. Nunca
escribe en el catalogo, y la exportacion es una propuesta, no un cambio.

EL ANCLAJE SE TRATA COMO UN RIESGO. La clasificacion automatica no se muestra de
entrada y exige un clic deliberado detras de una advertencia. Mostrar la
respuesta del clasificador al lado de la pregunta es como una referencia humana
se convierte en silencio en una copia de la maquina.
"""

import argparse
import collections
import csv
import hashlib
import json
from pathlib import Path

from multirag.paths import EXPERIMENTS_DIR


RECETA_VERSION = "interfaz-revision-v1"

SALIDA_PREDETERMINADA = EXPERIMENTS_DIR / "revision_catalogo_24"

CAMPOS_CORREGIBLES = (
    "emisor_id_corregido",
    "emisor_nombre_corregido",
    "tipo_documento_corregido",
    "dominios_documentales_corregidos",
)

# Which current field each correction replaces. The interface needs this to
# demand that a correction actually differ from what is on file.
# [ES] Que campo actual reemplaza cada correccion. La interfaz lo necesita para
# exigir que una correccion efectivamente difiera de lo que ya esta.
ORIGEN_DE_CORRECCION = {
    "emisor_id_corregido": "emisor_id",
    "emisor_nombre_corregido": "emisor_nombre",
    "tipo_documento_corregido": "tipo_documento",
    "dominios_documentales_corregidos": "dominios_documentales",
}

OPCIONES_DECISION = ("confirmar", "corregir", "dudoso", "excluir_del_corpus")


def huella_de_archivo(ruta: Path) -> str:
    h = hashlib.sha256()
    with ruta.open("rb") as f:
        for bloque in iter(lambda: f.read(1 << 20), b""):
            h.update(bloque)
    return f"sha256:{h.hexdigest()}"


def leer_csv_v2(ruta: Path):
    with ruta.open(encoding="utf-8-sig", newline="") as f:
        lector = csv.DictReader(f)
        return list(lector.fieldnames), list(lector)


def leer_defectos(ruta: Path) -> dict:
    """Per document: how many ingestion defects were registered.

    [ES] Por documento: cuantos defectos de ingesta quedaron registrados.
    """
    if not ruta.exists():
        return {}
    registros = json.loads(ruta.read_text(encoding="utf-8"))
    resumen = collections.defaultdict(lambda: {"registros": 0, "ocurrencias": 0})
    for r in registros:
        resumen[r["document_id"]]["registros"] += 1
        resumen[r["document_id"]]["ocurrencias"] += int(r.get("ocurrencias", 0))
    return dict(resumen)


def leer_contexto_automatico(ruta_xlsx: Path) -> dict:
    """The classifier's answer, read from the v2 workbook, kept behind a click.

    [ES] La respuesta del clasificador, leida del libro v2, guardada detras de un
    clic.
    """
    if not ruta_xlsx.exists():
        return {}
    from openpyxl import load_workbook

    libro = load_workbook(ruta_xlsx, read_only=True, data_only=True)
    if "contexto_automatico" not in libro.sheetnames:
        return {}
    hoja = libro["contexto_automatico"]
    filas = list(hoja.iter_rows(values_only=True))
    encabezado = None
    salida = {}
    for fila in filas:
        if fila and fila[0] == "document_id":
            encabezado = list(fila)
            continue
        if encabezado and fila and fila[0]:
            registro = dict(zip(encabezado, fila))
            salida[registro["document_id"]] = (
                registro.get("silos_de_chunks_persistidos") or ""
            )
    return salida


def incrustar_json(datos) -> str:
    """Embed JSON safely inside a script tag.

    `<` is escaped so no value can close the tag early. It is the one place
    where corpus text meets the document structure, and it is the one place a
    document could otherwise take over the page.

    [ES] Embebe JSON de forma segura dentro de una etiqueta script.

    Se escapa `<` para que ningun valor pueda cerrar la etiqueta antes de tiempo.
    Es el unico lugar donde el texto del corpus toca la estructura del documento,
    y el unico lugar donde un documento podria, si no, apoderarse de la pagina.
    """
    return json.dumps(datos, ensure_ascii=False).replace("<", "\\u003c")


PLANTILLA = r"""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Revisión del catálogo — 24 documentos</title>
<style>
:root{
  --tinta:#1c1c1e; --suave:#5b5f66; --linea:#dfe3e8; --fondo:#f5f6f8;
  --panel:#ffffff; --acento:#2f5597; --ok:#1e7a46; --corr:#8a5a00;
  --duda:#7a4fb5; --excl:#a32020; --aviso:#fff4e5; --avisoLinea:#e0a458;
}
*{box-sizing:border-box}
body{margin:0;background:var(--fondo);color:var(--tinta);
  font:15px/1.55 system-ui,-apple-system,"Segoe UI",Roboto,sans-serif}
header{background:var(--panel);border-bottom:1px solid var(--linea);
  padding:14px 22px;position:sticky;top:0;z-index:10}
.titulo{font-weight:650;font-size:16px}
.sub{color:var(--suave);font-size:13px;margin-top:2px}
.barra{height:8px;background:var(--linea);border-radius:99px;margin-top:12px;overflow:hidden}
.barra > div{height:100%;background:var(--acento);width:0;transition:width .18s}
.contadores{display:flex;gap:16px;flex-wrap:wrap;margin-top:10px;font-size:13px}
.contadores b{font-variant-numeric:tabular-nums}
main{max-width:1000px;margin:22px auto;padding:0 18px}
.ficha{background:var(--panel);border:1px solid var(--linea);border-radius:12px;padding:22px}
h1{font-size:21px;margin:0 0 2px}
.docid{color:var(--suave);font-size:13px;font-family:ui-monospace,Consolas,monospace}
.campos{display:grid;grid-template-columns:180px 1fr;gap:8px 16px;margin:18px 0}
.campos dt{color:var(--suave);font-size:13px}
.campos dd{margin:0}
.etiquetas{display:flex;gap:6px;flex-wrap:wrap}
.etiqueta{background:#eaf0fb;color:var(--acento);border:1px solid #cfdcf5;
  border-radius:99px;padding:2px 10px;font-size:13px}
.evidencia{background:#fafbfc;border:1px solid var(--linea);border-radius:8px;
  padding:12px 14px;margin:10px 0;white-space:pre-wrap;font-size:13.5px}
.evidencia .rotulo{color:var(--suave);font-size:12px;margin-bottom:6px;
  text-transform:uppercase;letter-spacing:.04em}
.aviso{background:var(--aviso);border:1px solid var(--avisoLinea);border-radius:8px;
  padding:11px 14px;margin:14px 0;font-size:13.5px}
.acciones{display:flex;gap:10px;flex-wrap:wrap;margin:20px 0 6px}
button{font:inherit;border:1px solid var(--linea);background:var(--panel);
  color:var(--tinta);border-radius:8px;padding:9px 16px;cursor:pointer}
button:hover{border-color:var(--suave)}
button:disabled{opacity:.45;cursor:not-allowed}
button.sel[data-d="confirmar"]{background:var(--ok);border-color:var(--ok);color:#fff}
button.sel[data-d="corregir"]{background:var(--corr);border-color:var(--corr);color:#fff}
button.sel[data-d="dudoso"]{background:var(--duda);border-color:var(--duda);color:#fff}
button.sel[data-d="excluir_del_corpus"]{background:var(--excl);border-color:var(--excl);color:#fff}
.correcciones{border:1px dashed var(--corr);border-radius:8px;padding:14px;margin-top:14px}
.correcciones label{display:block;margin-bottom:10px}
.correcciones span{display:block;color:var(--suave);font-size:12.5px;margin-bottom:3px}
input[type=text],textarea{width:100%;font:inherit;padding:8px 10px;
  border:1px solid var(--linea);border-radius:6px;background:#fff;color:var(--tinta)}
textarea{min-height:78px;resize:vertical}
.actual{color:var(--suave);font-size:12px;margin-top:2px}
.error{color:var(--excl);font-size:13px;margin-top:10px;min-height:19px}
.navegacion{display:flex;gap:10px;align-items:center;justify-content:space-between;
  margin-top:18px;flex-wrap:wrap}
select{font:inherit;padding:8px;border:1px solid var(--linea);border-radius:6px;background:#fff}
.pie{display:flex;gap:10px;flex-wrap:wrap;margin-top:22px;padding-top:18px;
  border-top:1px solid var(--linea)}
.peligro{border-color:var(--excl);color:var(--excl)}
.auto{margin-top:14px}
.auto .contenido{background:#fff6f6;border:1px solid #f0c7c7;border-radius:8px;
  padding:12px 14px;margin-top:8px;font-size:13.5px}
.oculto{display:none}
</style>
</head>
<body>
<header>
  <div class="titulo">Revisión humana del catálogo — 24 documentos canónicos</div>
  <div class="sub" id="sub"></div>
  <div class="barra"><div id="progreso"></div></div>
  <div class="contadores">
    <span>✓ confirmados <b id="c-confirmar">0</b></span>
    <span>✎ corregidos <b id="c-corregir">0</b></span>
    <span>? dudosos <b id="c-dudoso">0</b></span>
    <span>✕ excluidos <b id="c-excluir_del_corpus">0</b></span>
    <span>· pendientes <b id="c-pendientes">24</b></span>
  </div>
</header>

<main>
  <div id="aviso-persistencia" class="aviso oculto"></div>
  <section class="ficha">
    <div class="docid" id="posicion"></div>
    <h1 id="titulo"></h1>
    <div class="docid" id="docid"></div>

    <div id="aviso-defecto" class="aviso oculto"></div>

    <dl class="campos">
      <dt>Archivo</dt><dd id="archivo"></dd>
      <dt>Emisor actual</dt><dd id="emisor"></dd>
      <dt>Tipo documental actual</dt><dd id="tipo"></dd>
      <dt>Dominios actuales</dt><dd><div class="etiquetas" id="dominios"></div></dd>
      <dt>Jurisdicción y fecha</dt><dd id="jurisdiccion"></dd>
      <dt>Origen</dt><dd id="origen"></dd>
    </dl>

    <div class="evidencia"><div class="rotulo">Extracto de evidencia (textual)</div>
      <span id="extracto"></span></div>
    <div class="evidencia"><div class="rotulo">Evidencia por dominio — señala dónde mirar, no justifica la etiqueta</div>
      <span id="ev-dominio"></span></div>

    <div class="auto">
      <button type="button" id="btn-auto">Mostrar clasificación automática</button>
      <div class="contenido oculto" id="auto-contenido"></div>
    </div>

    <div class="acciones" id="acciones"></div>

    <div class="correcciones oculto" id="correcciones"></div>

    <label id="lbl-obs" class="oculto" style="display:block;margin-top:14px">
      <span style="display:block;color:#5b5f66;font-size:12.5px;margin-bottom:3px"
            id="txt-obs">Observaciones</span>
      <textarea id="observaciones"></textarea>
    </label>

    <div class="error" id="error"></div>

    <div class="navegacion">
      <div style="display:flex;gap:10px">
        <button type="button" id="btn-anterior">← Anterior</button>
        <button type="button" id="btn-siguiente">Siguiente →</button>
      </div>
      <select id="salto"></select>
    </div>
  </section>

  <div class="pie">
    <button type="button" id="btn-finalizar" disabled>Finalizar revisión</button>
    <button type="button" id="btn-csv" disabled>Descargar CSV</button>
    <button type="button" id="btn-json" disabled>Descargar JSON</button>
    <button type="button" id="btn-borrador">Descargar borrador (JSON)</button>
    <button type="button" id="btn-borrar" class="peligro">Borrar todas las decisiones</button>
  </div>
  <p class="sub" style="margin-top:14px">
    Todo se guarda solo en este navegador. Nada se envía a internet y nada se aplica
    al catálogo: la exportación es una propuesta para revisar.
  </p>
</main>

<script type="application/json" id="datos">__DATOS__</script>

<script>
/* === INICIO LOGICA PURA === */
/* Pure functions: no DOM, no storage. Extracted verbatim by the tests and run
   under node, so validation and export are actually tested and not merely
   inspected.
   [ES] Funciones puras: sin DOM, sin almacenamiento. Las pruebas las extraen tal
   cual y las corren con node, asi que la validacion y la exportacion se prueban
   de verdad y no solo se inspeccionan. */

function limpio(v){ return (v === undefined || v === null) ? "" : String(v).trim(); }

function hayCorreccionEfectiva(doc, estado, campos, origen){
  /* A correction has to differ from what is on file. Retyping the current value
     is not a correction, and accepting it would record a change that changes
     nothing.
     [ES] Una correccion tiene que diferir de lo que ya esta. Reescribir el valor
     actual no es una correccion, y aceptarlo registraria un cambio que no
     cambia nada. */
  for (const campo of campos){
    const nuevo = limpio(estado[campo]);
    if (nuevo !== "" && nuevo !== limpio(doc[origen[campo]])) return true;
  }
  return false;
}

function validarDecision(doc, estado, campos, origen){
  const d = limpio(estado.decision_humana);
  if (d === "") return {ok:false, motivo:"Sin decisión."};

  if (d === "confirmar"){
    for (const campo of campos){
      if (limpio(estado[campo]) !== ""){
        return {ok:false, motivo:"«Confirmar» no admite correcciones cargadas."};
      }
    }
    return {ok:true, motivo:""};
  }
  if (d === "corregir"){
    if (!hayCorreccionEfectiva(doc, estado, campos, origen)){
      return {ok:false,
        motivo:"Completá al menos un campo con un valor distinto del actual."};
    }
    return {ok:true, motivo:""};
  }
  if (d === "dudoso"){
    if (limpio(estado.observaciones) === ""){
      return {ok:false, motivo:"«Dudoso» exige explicar la duda en observaciones."};
    }
    return {ok:true, motivo:""};
  }
  if (d === "excluir_del_corpus"){
    if (limpio(estado.observaciones) === ""){
      return {ok:false, motivo:"«Excluir del corpus» exige justificar el motivo."};
    }
    return {ok:true, motivo:""};
  }
  return {ok:false, motivo:"Decisión desconocida."};
}

function contarDecisiones(documentos, estados, opciones, campos, origen){
  const cuenta = {pendientes:0};
  for (const o of opciones) cuenta[o] = 0;
  for (const doc of documentos){
    const estado = estados[doc.document_id] || {};
    const v = validarDecision(doc, estado, campos, origen);
    if (v.ok) cuenta[limpio(estado.decision_humana)] += 1;
    else cuenta.pendientes += 1;
  }
  return cuenta;
}

function celdaCsv(valor){
  const v = (valor === undefined || valor === null) ? "" : String(valor);
  return /[",\n\r]/.test(v) ? '"' + v.replace(/"/g, '""') + '"' : v;
}

function armarCsv(documentos, estados, columnas, humanas){
  /* The 22 columns of the v2 sheet, in the same order, with the six human ones
     filled in. Anything the reviewer left empty stays empty: empty means "I did
     not touch it", never "delete it".
     [ES] Las 22 columnas de la planilla v2, en el mismo orden, con las seis
     humanas completas. Lo que el revisor dejo vacio queda vacio: vacio significa
     "no lo toque", nunca "borralo". */
  const lineas = [columnas.map(celdaCsv).join(",")];
  for (const doc of documentos){
    const estado = estados[doc.document_id] || {};
    const fila = columnas.map(function(col){
      return celdaCsv(humanas.indexOf(col) >= 0 ? limpio(estado[col]) : doc[col]);
    });
    lineas.push(fila.join(","));
  }
  return lineas.join("\n") + "\n";
}

function armarJson(meta, documentos, estados, opciones, campos, origen, humanas){
  const decisiones = documentos.map(function(doc){
    const estado = estados[doc.document_id] || {};
    const registro = {document_id: doc.document_id, fuente: doc.fuente};
    for (const col of humanas) registro[col] = limpio(estado[col]);
    registro.valida = validarDecision(doc, estado, campos, origen).ok;
    return registro;
  });
  return {
    receta: meta.receta,
    huella_catalogo: meta.huella_catalogo,
    huella_csv_v2: meta.huella_csv_v2,
    fecha_exportacion: meta.fecha_exportacion,
    documentos: documentos.length,
    conteos: contarDecisiones(documentos, estados, opciones, campos, origen),
    decisiones: decisiones,
    salvedades: meta.salvedades
  };
}
/* === FIN LOGICA PURA === */

(function(){
  "use strict";
  const D = JSON.parse(document.getElementById("datos").textContent);
  const CAMPOS = D.campos_corregibles;
  const ORIGEN = D.origen_de_correccion;
  const HUMANAS = D.columnas_humanas;
  const CLAVE = "revision_catalogo_24::" + D.huella_csv_v2.slice(0, 24);

  let indice = 0;
  let estados = {};
  let persistencia = true;

  /* Probe storage instead of assuming it. It is disabled in private windows,
     under some file:// configurations and inside data: URLs. Swallowing the
     failure silently is how a reviewer loses twenty judgements and finds out
     at the end.
     [ES] Se sondea el almacenamiento en lugar de suponerlo. Esta deshabilitado
     en ventanas privadas, en algunas configuraciones de file:// y dentro de
     URLs data:. Tragarse la falla en silencio es como un revisor pierde veinte
     juicios y se entera al final. */
  try {
    const sonda = CLAVE + "::sonda";
    localStorage.setItem(sonda, "1");
    localStorage.removeItem(sonda);
  } catch (e) { persistencia = false; }

  if (persistencia){
    try { estados = JSON.parse(localStorage.getItem(CLAVE) || "{}"); } catch (e) { estados = {}; }
  }

  function guardar(){
    if (!persistencia) return;
    try { localStorage.setItem(CLAVE, JSON.stringify(estados)); }
    catch (e) { persistencia = false; avisarPersistencia(); }
  }

  function avisarPersistencia(){
    const caja = document.getElementById("aviso-persistencia");
    if (persistencia){ caja.classList.add("oculto"); return; }
    caja.classList.remove("oculto");
    caja.textContent = "⚠ Este navegador no permite guardado automático acá "
      + "(pasa en ventanas privadas y en algunos modos de apertura). Tus "
      + "decisiones NO se conservan si cerrás la pestaña. Usá «Descargar "
      + "borrador (JSON)» seguido para no perder trabajo.";
  }
  function estadoDe(doc){
    if (!estados[doc.document_id]) estados[doc.document_id] = {};
    return estados[doc.document_id];
  }
  function texto(id, valor){
    document.getElementById(id).textContent = (valor === null || valor === undefined || valor === "") ? "—" : String(valor);
  }
  function vaciar(nodo){ while (nodo.firstChild) nodo.removeChild(nodo.firstChild); }

  function pintarDominios(doc){
    const cont = document.getElementById("dominios");
    vaciar(cont);
    const tokens = String(doc.dominios_documentales || "").split("|")
      .map(function(t){ return t.trim(); }).filter(Boolean);
    if (!tokens.length){
      const s = document.createElement("span"); s.textContent = "—"; cont.appendChild(s); return;
    }
    tokens.forEach(function(t){
      const s = document.createElement("span");
      s.className = "etiqueta";
      s.textContent = t;              // textContent: viene de un documento
      cont.appendChild(s);
    });
  }

  function pintarOrigen(doc){
    const cont = document.getElementById("origen");
    vaciar(cont);
    const url = String(doc.url_origen || "").trim();
    /* Only http/https become clickable. A `javascript:` value in a catalogue
       field must never turn into a link.
       [ES] Solo http/https se vuelven clickeables. Un valor `javascript:` en un
       campo del catalogo no puede convertirse nunca en un enlace. */
    if (/^https?:\/\//i.test(url)){
      const a = document.createElement("a");
      a.href = url; a.textContent = url;
      a.target = "_blank"; a.rel = "noreferrer noopener";
      cont.appendChild(a);
    } else {
      const s = document.createElement("span");
      s.textContent = url || "—";
      cont.appendChild(s);
    }
  }

  function pintarAcciones(doc){
    const cont = document.getElementById("acciones");
    vaciar(cont);
    const estado = estadoDe(doc);
    D.opciones_decision.forEach(function(op){
      const b = document.createElement("button");
      b.type = "button";
      b.dataset.d = op;
      b.textContent = D.etiquetas_decision[op];
      if (estado.decision_humana === op) b.className = "sel";
      b.addEventListener("click", function(){
        estado.decision_humana = op;
        if (op === "confirmar") CAMPOS.forEach(function(c){ estado[c] = ""; });
        guardar(); render();
      });
      cont.appendChild(b);
    });
  }

  function pintarCorrecciones(doc){
    const cont = document.getElementById("correcciones");
    vaciar(cont);
    const estado = estadoDe(doc);
    if (estado.decision_humana !== "corregir"){ cont.classList.add("oculto"); return; }
    cont.classList.remove("oculto");

    const aviso = document.createElement("div");
    aviso.className = "actual";
    aviso.textContent = "Completá solo lo que cambia. Un campo vacío conserva el valor actual.";
    cont.appendChild(aviso);

    CAMPOS.forEach(function(campo){
      const label = document.createElement("label");
      const span = document.createElement("span");
      span.textContent = campo;
      const input = document.createElement("input");
      input.type = "text";
      input.value = estado[campo] || "";
      input.addEventListener("input", function(){
        estado[campo] = input.value; guardar(); actualizarCabecera(); validar(doc);
      });
      const actual = document.createElement("div");
      actual.className = "actual";
      actual.textContent = "actual: " + (doc[ORIGEN[campo]] || "—");
      label.appendChild(span); label.appendChild(input); label.appendChild(actual);
      cont.appendChild(label);
    });
  }

  function pintarObservaciones(doc){
    const estado = estadoDe(doc);
    const label = document.getElementById("lbl-obs");
    const area = document.getElementById("observaciones");
    const rotulo = document.getElementById("txt-obs");
    const d = estado.decision_humana;
    label.classList.remove("oculto");
    if (d === "dudoso") rotulo.textContent = "Observaciones — obligatorio: explicá la duda";
    else if (d === "excluir_del_corpus") rotulo.textContent = "Observaciones — obligatorio: justificá el motivo";
    else rotulo.textContent = "Observaciones (opcional)";
    area.value = estado.observaciones || "";
    area.oninput = function(){
      estado.observaciones = area.value; guardar(); actualizarCabecera(); validar(doc);
    };
  }

  function pintarDefecto(doc){
    const caja = document.getElementById("aviso-defecto");
    const defecto = D.defectos[doc.document_id];
    if (!defecto){ caja.classList.add("oculto"); caja.textContent = ""; return; }
    caja.classList.remove("oculto");
    caja.textContent = "⚠ Este documento contiene caracteres dañados durante la "
      + "ingesta. El defecto está registrado aparte y no debe corregirse desde "
      + "esta pantalla. (" + defecto.ocurrencias + " ocurrencias en "
      + defecto.registros + " fragmentos)";
  }

  function pintarAutomatico(doc){
    const caja = document.getElementById("auto-contenido");
    const boton = document.getElementById("btn-auto");
    vaciar(caja);
    caja.classList.add("oculto");
    boton.textContent = "Mostrar clasificación automática";
    boton.onclick = function(){
      if (caja.classList.contains("oculto")){
        vaciar(caja);
        const p1 = document.createElement("div");
        p1.textContent = "⚠ No es verdad de referencia: es lo que decidió el "
          + "clasificador. Mirarlo antes de decidir puede anclar tu juicio hacia "
          + "coincidir con la máquina.";
        const p2 = document.createElement("div");
        p2.style.marginTop = "8px";
        p2.textContent = D.contexto_automatico[doc.document_id] || "(sin dato)";
        caja.appendChild(p1); caja.appendChild(p2);
        caja.classList.remove("oculto");
        boton.textContent = "Ocultar clasificación automática";
      } else {
        caja.classList.add("oculto");
        boton.textContent = "Mostrar clasificación automática";
      }
    };
  }

  function pintarSalto(){
    const sel = document.getElementById("salto");
    vaciar(sel);
    D.documentos.forEach(function(doc, i){
      const estado = estados[doc.document_id] || {};
      const v = validarDecision(doc, estado, CAMPOS, ORIGEN);
      const marca = v.ok ? D.marcas[estado.decision_humana] : "·";
      const op = document.createElement("option");
      op.value = String(i);
      op.textContent = marca + "  " + (i + 1) + "/" + D.documentos.length + "  " + doc.document_id + " — " + doc.fuente;
      if (i === indice) op.selected = true;
      sel.appendChild(op);
    });
    sel.onchange = function(){ indice = parseInt(sel.value, 10); render(); };
  }

  function validar(doc){
    const v = validarDecision(doc, estadoDe(doc), CAMPOS, ORIGEN);
    document.getElementById("error").textContent =
      (estadoDe(doc).decision_humana && !v.ok) ? v.motivo : "";
    return v;
  }

  function actualizarCabecera(){
    const c = contarDecisiones(D.documentos, estados, D.opciones_decision, CAMPOS, ORIGEN);
    D.opciones_decision.forEach(function(op){
      document.getElementById("c-" + op).textContent = String(c[op]);
    });
    document.getElementById("c-pendientes").textContent = String(c.pendientes);
    const hechos = D.documentos.length - c.pendientes;
    document.getElementById("progreso").style.width =
      (hechos / D.documentos.length * 100).toFixed(1) + "%";
    document.getElementById("sub").textContent =
      hechos + " de " + D.documentos.length + " con decisión válida";
    const completo = c.pendientes === 0;
    ["btn-finalizar", "btn-csv", "btn-json"].forEach(function(id){
      document.getElementById(id).disabled = !completo;
    });
    pintarSalto();
  }

  function render(){
    const doc = D.documentos[indice];
    texto("posicion", "Documento " + (indice + 1) + " de " + D.documentos.length);
    texto("titulo", doc.titulo_oficial || doc.fuente);
    texto("docid", doc.document_id + "  ·  " + doc.fuente);
    texto("archivo", doc.archivo_referencia);
    texto("emisor", (doc.emisor_nombre || "—") + "  (" + (doc.emisor_id || "—") + ")");
    texto("tipo", doc.tipo_documento);
    texto("jurisdiccion", (doc.jurisdiccion || "—") + "  ·  " + (doc.fecha_documento || "—"));
    texto("extracto", doc.evidencia_extracto);
    texto("ev-dominio", doc.evidencia_por_dominio);
    pintarDominios(doc);
    pintarOrigen(doc);
    pintarDefecto(doc);
    pintarAutomatico(doc);
    pintarAcciones(doc);
    pintarCorrecciones(doc);
    pintarObservaciones(doc);
    validar(doc);
    document.getElementById("btn-anterior").disabled = indice === 0;
    document.getElementById("btn-siguiente").disabled = indice === D.documentos.length - 1;
    actualizarCabecera();
    window.scrollTo(0, 0);
  }

  function descargar(nombre, contenido, tipo){
    const blob = new Blob([contenido], {type: tipo});
    const url = URL.createObjectURL(blob);
    const a = document.createElement("a");
    a.href = url; a.download = nombre;
    document.body.appendChild(a); a.click();
    document.body.removeChild(a); URL.revokeObjectURL(url);
  }

  function meta(){
    return {
      receta: D.receta,
      huella_catalogo: D.huella_catalogo,
      huella_csv_v2: D.huella_csv_v2,
      fecha_exportacion: new Date().toISOString(),
      salvedades: D.salvedades
    };
  }

  document.getElementById("btn-anterior").onclick = function(){
    if (indice > 0){ indice -= 1; render(); }
  };
  document.getElementById("btn-siguiente").onclick = function(){
    if (indice < D.documentos.length - 1){ indice += 1; render(); }
  };
  document.getElementById("btn-csv").onclick = function(){
    descargar("revision_catalogo_24_decisiones.csv",
      "﻿" + armarCsv(D.documentos, estados, D.columnas, HUMANAS),
      "text/csv;charset=utf-8");
  };
  document.getElementById("btn-json").onclick = function(){
    descargar("revision_catalogo_24_decisiones.json",
      JSON.stringify(armarJson(meta(), D.documentos, estados,
        D.opciones_decision, CAMPOS, ORIGEN, HUMANAS), null, 2),
      "application/json;charset=utf-8");
  };
  document.getElementById("btn-borrador").onclick = function(){
    /* Available at any time, complete or not. Partial work must never be
       trapped inside a tab.
       [ES] Disponible en cualquier momento, completo o no. El trabajo parcial
       nunca puede quedar atrapado dentro de una pestana. */
    descargar("revision_catalogo_24_borrador.json",
      JSON.stringify(armarJson(meta(), D.documentos, estados,
        D.opciones_decision, CAMPOS, ORIGEN, HUMANAS), null, 2),
      "application/json;charset=utf-8");
  };
  document.getElementById("btn-finalizar").onclick = function(){
    const c = contarDecisiones(D.documentos, estados, D.opciones_decision, CAMPOS, ORIGEN);
    window.alert("Revisión completa: " + (D.documentos.length - c.pendientes)
      + " de " + D.documentos.length + " documentos.\n\n"
      + "confirmados " + c.confirmar + "  ·  corregidos " + c.corregir
      + "  ·  dudosos " + c.dudoso + "  ·  excluidos " + c.excluir_del_corpus
      + "\n\nDescargá el CSV y el JSON. Nada se aplicó al catálogo.");
  };
  document.getElementById("btn-borrar").onclick = function(){
    if (!window.confirm("¿Borrar TODAS las decisiones cargadas? No se puede deshacer.")) return;
    if (!window.confirm("Confirmá otra vez: se pierden las 24 decisiones.")) return;
    estados = {};
    try { localStorage.removeItem(CLAVE); } catch (e) {}
    indice = 0; render();
  };

  avisarPersistencia();
  render();
})();
</script>
</body>
</html>
"""


def construir_html(datos) -> str:
    return PLANTILLA.replace("__DATOS__", incrustar_json(datos))


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--carpeta", type=Path, default=SALIDA_PREDETERMINADA)
    args = parser.parse_args()

    csv_v2 = args.carpeta / "revision_catalogo_24_v2.csv"
    xlsx_v2 = args.carpeta / "revision_catalogo_24_v2.xlsx"
    defectos_json = args.carpeta / "defectos_de_ingesta.json"
    manifest_v2 = args.carpeta / "manifest_v2.json"
    salida = args.carpeta / "revision_catalogo_24.html"

    if not csv_v2.exists():
        raise SystemExit(f"no existe la fuente: {csv_v2}")

    columnas, documentos = leer_csv_v2(csv_v2)
    manifest = json.loads(manifest_v2.read_text(encoding="utf-8"))

    datos = {
        "receta": RECETA_VERSION,
        "huella_catalogo": manifest["huella_catalogo"],
        "huella_csv_v2": huella_de_archivo(csv_v2),
        "columnas": columnas,
        "columnas_humanas": list(manifest["columnas_a_completar"]),
        "campos_corregibles": list(CAMPOS_CORREGIBLES),
        "origen_de_correccion": dict(ORIGEN_DE_CORRECCION),
        "opciones_decision": list(OPCIONES_DECISION),
        "etiquetas_decision": {
            "confirmar": "✓ Confirmar",
            "corregir": "✎ Corregir",
            "dudoso": "? Dudoso",
            "excluir_del_corpus": "✕ Excluir del corpus",
        },
        "marcas": {
            "confirmar": "✓", "corregir": "✎",
            "dudoso": "?", "excluir_del_corpus": "✕",
        },
        "documentos": documentos,
        "defectos": leer_defectos(defectos_json),
        "contexto_automatico": leer_contexto_automatico(xlsx_v2),
        "salvedades": [
            "Los metadatos del catalogo estaban PENDIENTES DE RATIFICACION HUMANA "
            "al momento de esta revision.",
            "La exportacion es una PROPUESTA. No se aplico nada al catalogo ni a "
            "PostgreSQL.",
            "Un campo de correccion vacio conserva el valor actual: vacio significa "
            "'no lo toque', nunca 'borralo'.",
            "Los caracteres danados de la ingesta se registran aparte y no se "
            "corrigen desde esta pantalla.",
            "Si el navegador no permite guardado automatico, la interfaz lo avisa "
            "en pantalla y el borrador se puede descargar en cualquier momento.",
        ],
    }

    salida.write_text(construir_html(datos), encoding="utf-8")

    print(f"documentos    {len(documentos)}")
    print(f"columnas      {len(columnas)}")
    print(f"defectos      {len(datos['defectos'])} documento(s) marcado(s)")
    print(f"tamano        {salida.stat().st_size / 1024:.0f} KB")
    print()
    print(f"html          {salida}")


if __name__ == "__main__":
    main()
