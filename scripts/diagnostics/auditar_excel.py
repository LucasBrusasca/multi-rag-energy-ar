"""Diagnostic audit of a spreadsheet: native cells versus what Docling produces.

Spreadsheets have a separate path in the audit because they need no OCR: the
cells, their merges and their formulas are already structured data in the file.
Anything lost there was lost by the reader, never by a recognition step.

The workbook audited lives in `data/quarantine/descartados/`: it was excluded
from the corpus and never ingested. It is audited anyway because it is the only
spreadsheet in the project and it carries exactly the case the audit asks for —
a hierarchical header where the period spans several columns and the unit sits
in a cell of its own.

STRICTLY READ-ONLY.

[ES] Auditoría diagnóstica de una planilla: celdas nativas contra lo que produce
Docling.

Las planillas tienen una vía separada en la auditoría porque no necesitan OCR:
las celdas, sus combinaciones y sus fórmulas ya son datos estructurados dentro
del archivo. Todo lo que se pierda ahí lo perdió el lector, nunca un paso de
reconocimiento.

El libro auditado vive en `data/quarantine/descartados/`: fue excluido del
corpus y nunca se ingirió. Se audita igual porque es la única planilla del
proyecto y trae exactamente el caso que la auditoría pide — un encabezado
jerárquico donde el período abarca varias columnas y la unidad está en una celda
propia.

ESTRICTAMENTE DE SOLO LECTURA.
"""

import argparse
import collections
from pathlib import Path

from multirag.paths import DATA_DIR


PLANILLA_PREDETERMINADA = (
    DATA_DIR / "quarantine" / "descartados" / "1Q26.xlsx"
)


def leer_nativo(ruta: Path, hojas=None) -> list[dict]:
    """Read the sheets as the file really holds them.

    [ES] Lee las hojas tal como el archivo las guarda realmente.
    """
    import openpyxl

    libro = openpyxl.load_workbook(ruta, data_only=False)
    elegidas = hojas or libro.sheetnames

    informe = []

    for nombre in elegidas:
        if nombre not in libro.sheetnames:
            continue

        hoja = libro[nombre]

        formulas = sum(
            1
            for fila in hoja.iter_rows()
            for celda in fila
            if isinstance(celda.value, str) and celda.value.startswith("=")
        )

        # Fully empty columns act as visual spacers and are the usual cause of
        # a reader splitting one sheet into several tables.
        # [ES] Las columnas totalmente vacías funcionan como separadores
        # visuales y son la causa habitual de que un lector parta una hoja en
        # varias tablas.
        columnas_vacias = [
            indice
            for indice in range(1, hoja.max_column + 1)
            if all(
                hoja.cell(row=f, column=indice).value in (None, "")
                for f in range(1, min(hoja.max_row, 60) + 1)
            )
        ]

        informe.append(
            {
                "hoja": nombre,
                "filas": hoja.max_row,
                "columnas": hoja.max_column,
                "combinadas": len(hoja.merged_cells.ranges),
                "rangos_combinados": [
                    str(r) for r in list(hoja.merged_cells.ranges)[:8]
                ],
                "formulas": formulas,
                "columnas_vacias": columnas_vacias,
                "cabecera": [
                    [
                        (celda.coordinate, celda.value)
                        for celda in fila
                        if celda.value is not None
                    ]
                    for fila in hoja.iter_rows(min_row=1, max_row=4)
                ],
            }
        )

    return informe


def leer_docling(ruta: Path) -> dict:
    """Read the same file through Docling and describe how it split it.

    [ES] Lee el mismo archivo con Docling y describe cómo lo partió.
    """
    from docling.document_converter import DocumentConverter

    resultado = DocumentConverter().convert(str(ruta), raises_on_error=False)

    tablas = resultado.document.tables

    dimensiones = collections.Counter(
        f"{t.data.num_rows}x{t.data.num_cols}" for t in tablas
    )

    return {
        "estado": resultado.status.name,
        "tablas": len(tablas),
        "textos": len(resultado.document.texts),
        "una_columna": sum(1 for t in tablas if t.data.num_cols == 1),
        "una_celda": sum(
            1
            for t in tablas
            if t.data.num_rows == 1 and t.data.num_cols == 1
        ),
        "dimensiones": dimensiones.most_common(8),
        "muestra": [
            {
                "self_ref": t.self_ref,
                "dimensiones": f"{t.data.num_rows}x{t.data.num_cols}",
                "celdas": [c.text[:24] for c in t.data.table_cells[:6]],
            }
            for t in tablas[30:40]
        ],
    }


def construir_parser() -> argparse.ArgumentParser:
    """Build the command-line interface.

    [ES] Construye la interfaz de línea de comandos.
    """
    parser = argparse.ArgumentParser(
        description=(
            "Audita una planilla: celdas nativas contra la lectura de Docling. "
            "Solo lectura."
        )
    )
    parser.add_argument(
        "--planilla",
        type=Path,
        default=PLANILLA_PREDETERMINADA,
        help="Archivo .xlsx a auditar.",
    )
    parser.add_argument(
        "--hojas",
        nargs="*",
        default=["BCE ing", "EERR-C ing"],
        help="Hojas a inspeccionar de forma nativa.",
    )
    return parser


def main() -> None:
    """Run the spreadsheet audit.

    [ES] Corre la auditoría de la planilla.
    """
    argumentos = construir_parser().parse_args()
    ruta = Path(argumentos.planilla).resolve()

    if not ruta.is_file():
        raise SystemExit(f"No existe la planilla: {ruta}")

    print(f"planilla: {ruta}")
    print(
        "estado en el proyecto: "
        + (
            "EN CUARENTENA, nunca ingerida"
            if "quarantine" in str(ruta)
            else "en el corpus"
        )
    )

    print("\n=== 1. Celdas nativas (sin OCR: el archivo ya es estructurado) ===")

    for hoja in leer_nativo(ruta, argumentos.hojas):
        print(f"\nhoja {hoja['hoja']!r}: "
              f"{hoja['filas']} filas x {hoja['columnas']} columnas")
        print(f"  celdas combinadas : {hoja['combinadas']} "
              f"{hoja['rangos_combinados']}")
        print(f"  formulas          : {hoja['formulas']}")
        print(f"  columnas vacias   : {hoja['columnas_vacias']} "
              "(separadores visuales)")
        print("  encabezado:")
        for fila in hoja["cabecera"]:
            if fila:
                print(f"    {fila}")

    print("\n=== 2. Lectura de Docling sobre el mismo archivo ===")

    docling = leer_docling(ruta)

    print(f"estado           : {docling['estado']}")
    print(f"tablas producidas: {docling['tablas']}")
    print(f"  de una sola columna: {docling['una_columna']} "
          f"({docling['una_columna'] / max(1, docling['tablas']):.0%})")
    print(f"  de una sola celda  : {docling['una_celda']}")
    print(f"dimensiones mas frecuentes: {docling['dimensiones']}")
    print("\nmuestra de tablas consecutivas:")

    for t in docling["muestra"]:
        print(f"  {t['self_ref']:<16} {t['dimensiones']:>6}  {t['celdas']}")


if __name__ == "__main__":
    main()
