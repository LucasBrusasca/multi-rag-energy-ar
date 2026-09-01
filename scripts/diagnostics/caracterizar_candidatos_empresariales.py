"""FASE 3 v2 - Characterise the corporate candidates by READING them.

WHY v1 WAS NOT ENOUGH. v1 proposed a document type from the link text and a
period from whatever four digits appeared first - which included the year in the
URL directory. It then assigned every corporate document to BOTH `contable` and
`financiero` by fiat. Three of those are guesses dressed as data, and this
version replaces them with evidence read out of the file.

WHAT CHANGES:

  PERIOD, BY PRIORITY OF SOURCE. Content and PDF metadata first, then the file
  name, then the link text. The year in the URL directory is NEVER used on its
  own: `/uploads/2026/07/EEFF-31-12-25.pdf` is a 2025 statement published in
  2026, and taking the directory would date it wrong by a year. Where the
  sources disagree, the disagreement is reported.

  DOMAIN, ONLY WITH EVIDENCE. Each of `legal`, `impositivo`, `contable` and
  `financiero` is proposed only when enough distinct domain terms appear in the
  text, and every proposal carries the sentence and the page where it was found.
  A financial statement that mentions income tax once is not a tax document, and
  the threshold is what keeps it from being labelled as one.

  RELEVANCE. Terms and conditions, codes of ethics and sustainability reports
  are flagged as candidates for exclusion: they are corporate documents that
  carry almost no regulated-domain content, and padding the corpus with them
  would inflate a count without adding evidence.

  DOCUMENTARY DEDUPLICATION, beyond SHA-256: entity, type, period, normalised
  title, page count and a text fingerprint. Two renderings of the same statement
  have different bytes and the same content.

IT PROPOSES. It does not decide, does not ingest and does not touch the
catalogue. Every record leaves as `pendiente_revision`.

[ES] FASE 3 v2 - Caracterizar los candidatos empresariales LEYENDOLOS.

POR QUE LA v1 NO ALCANZABA. La v1 proponia un tipo documental a partir del texto
del enlace y un periodo a partir de los primeros cuatro digitos que aparecieran,
que incluian el ano del directorio de la URL. Despues asignaba por decreto todo
documento empresarial a `contable` Y `financiero`. Tres de esas cosas son
conjeturas disfrazadas de dato, y esta version las reemplaza por evidencia leida
del archivo.

QUE CAMBIA:

  PERIODO, POR PRIORIDAD DE FUENTE. Primero contenido y metadatos del PDF,
  despues el nombre de archivo, despues el texto del enlace. El ano del
  directorio de la URL NUNCA se usa solo: `/uploads/2026/07/EEFF-31-12-25.pdf` es
  un estado de 2025 publicado en 2026, y tomar el directorio lo fecharia mal por
  un ano. Donde las fuentes discrepan, la discrepancia se reporta.

  DOMINIO, SOLO CON EVIDENCIA. Cada uno de `legal`, `impositivo`, `contable` y
  `financiero` se propone solo cuando aparecen suficientes terminos distintos del
  dominio en el texto, y toda propuesta lleva la oracion y la pagina donde se
  encontro. Un estado financiero que menciona el impuesto a las ganancias una vez
  no es un documento impositivo, y el umbral es lo que impide etiquetarlo asi.

  PERTINENCIA. Terminos y condiciones, codigos de etica y reportes de
  sostenibilidad se marcan como candidatos a exclusion: son documentos
  corporativos con casi nada de contenido de dominio regulado, y rellenar el
  corpus con ellos inflaria un conteo sin agregar evidencia.

  DEDUPLICACION DOCUMENTAL, mas alla del SHA-256: entidad, tipo, periodo, titulo
  normalizado, cantidad de paginas y huella de texto. Dos renderizados del mismo
  estado tienen distintos bytes y el mismo contenido.

PROPONE. No decide, no ingiere y no toca el catalogo. Todo registro sale como
`pendiente_revision`.
"""

import argparse
import collections
import csv
import hashlib
import json
import re
import unicodedata
from pathlib import Path

from multirag.paths import DATA_DIR, PROJECT_ROOT


RECETA_VERSION = "caracterizacion-fase3v2-v1"

CANDIDATOS = DATA_DIR / "incoming" / "candidates"
CUARENTENA = DATA_DIR / "quarantine" / "descartados"
ADQUISICION = DATA_DIR / "catalog" / "candidates" / "adquisicion_fase3.json"
CATALOGO = DATA_DIR / "catalog" / "metadatos_curados.csv"
RAW = DATA_DIR / "raw"

PAGINAS_A_LEER = 25          # cabeza del documento
PAGINAS_FINALES = 3          # cola, donde suelen ir firmas y notas

# A domain is proposed only when at least this many DISTINCT terms of that
# domain appear. One mention of "impuesto a las ganancias" inside a balance
# sheet does not make the document a tax document.
# [ES] Un dominio se propone solo si aparecen al menos esta cantidad de terminos
# DISTINTOS de ese dominio. Una mencion de "impuesto a las ganancias" dentro de
# un balance no vuelve impositivo al documento.
# A domain is proposed only when THREE conditions hold at once. One term
# appearing once on one page is an incidental mention, not subject matter.
# [ES] Un dominio se propone solo si se cumplen TRES condiciones a la vez. Un
# termino que aparece una vez en una pagina es una mencion incidental, no
# materia.
UMBRAL_TERMINOS = 3          # terminos distintos del dominio
UMBRAL_OCURRENCIAS = 6       # menciones totales
UMBRAL_PAGINAS = 2           # algun termino en al menos dos paginas distintas

# Terms are matched with LEXICAL BOUNDARIES, not as substrings. Substring
# matching produced false positives that emptied whole domains of meaning:
#   `iva`        matched `comparativa`, `derivada`, `activa`, `positiva`
#   `arca`       matched `abarca`, `marca`, `demarca`
#   `percepcion` matched `percepcion de corrupcion`, which is not a tax regime
# Ambiguous single words were replaced by the collocations that actually name
# the concept. `resolucion` and `decreto` were dropped from the legal list
# entirely: they appear in nearly every Argentine document and would have made
# every document legal.
#
# [ES] Los terminos se buscan con LIMITES LEXICOS, no como subcadenas. La
# busqueda por subcadena producia falsos positivos que vaciaban de sentido a
# dominios enteros:
#   `iva`        matcheaba `comparativa`, `derivada`, `activa`, `positiva`
#   `arca`       matcheaba `abarca`, `marca`, `demarca`
#   `percepcion` matcheaba `percepcion de corrupcion`, que no es un regimen
#                impositivo
# Las palabras sueltas ambiguas se reemplazaron por las colocaciones que
# efectivamente nombran el concepto. `resolucion` y `decreto` se sacaron por
# completo de la lista legal: aparecen en casi todo documento argentino y
# habrian vuelto legal a todos.
TERMINOS = {
    "contable": [
        "estado de situacion patrimonial", "estados financieros", "estados contables",
        "patrimonio neto", "activo corriente", "pasivo corriente",
        "estado de resultados", "flujo de efectivo", "niif", "facpce",
        "informe del auditor", "notas a los estados",
        "balance sheet", "statement of financial position", "financial statements",
        "shareholders equity", "current assets", "current liabilities",
        "income statement", "cash flow", "ifrs", "independent auditor",
        "depreciation and amortization",
    ],
    "financiero": [
        "obligaciones negociables", "calificacion de riesgo", "ebitda",
        "deuda financiera", "tasa de interes", "emision de deuda",
        "estructura de capital", "apalancamiento", "prospecto",
        "suplemento de precio", "valor nominal",
        "net debt", "financial debt", "corporate bonds", "notes due",
        "interest rate", "credit rating", "leverage", "capital structure",
        "principal amount",
    ],
    "impositivo": [
        "impuesto a las ganancias", "impuesto al valor agregado", "iva",
        "afip", "arca", "alicuota", "declaracion jurada", "ingresos brutos",
        "base imponible", "credito fiscal", "impuesto diferido",
        "agente de retencion", "regimen de retencion", "retenciones impositivas",
        "agente de percepcion", "regimen de percepcion",
        "income tax", "value added tax", "deferred tax", "tax credit",
        "taxable income", "turnover tax", "withholding tax",
    ],
    "legal": [
        "enre", "enargas", "secretaria de energia", "marco regulatorio",
        "concesion", "licencia de distribucion", "cuadro tarifario",
        "audiencia publica", "ente regulador", "servicio publico",
        "regulatory framework", "public hearing", "regulatory authority",
        "secretariat of energy",
    ],
}

# Documents whose content is corporate but carries almost no regulated-domain
# material. Flagged, not deleted: the human review decides.
# [ES] Documentos de contenido corporativo pero con casi nada de materia
# regulada. Se marcan, no se borran: decide la revision humana.
PATRONES_NO_PERTINENTES = [
    ("terminos_y_condiciones", r"t[eé]rminos y condiciones|tyc[_\s-]|adhesi[oó]n al servicio"),
    ("codigo_de_etica", r"c[oó]digo de [eé]tica|code of (ethics|conduct)"),
    ("sostenibilidad", r"sustentabilidad|sostenibilidad|reporte asg|\besg\b|sustainability"),
    ("aviso_o_convocatoria", r"aviso de|convocatoria|anuncio_?fecha"),
]

# Order matters: the FIRST match wins, so the more specific type goes first.
# `memoria_anual` before `estado_financiero`, because an annual report contains
# financial statements while a set of statements is not an annual report.
# [ES] El orden importa: gana la PRIMERA coincidencia, asi que el tipo mas
# especifico va antes. `memoria_anual` antes que `estado_financiero`, porque una
# memoria contiene estados financieros mientras que un juego de estados no es
# una memoria.
TIPOS = [
    ("memoria_anual", r"\bmemoria\b|annual report|form 20-f|20-?f\b"),
    ("estado_financiero", r"estados? (financieros?|contables?)|financial statements|eeff"),
    ("reporte_resultados", r"earnings|resultados del (trimestre|periodo)|results release|earnings release"),
    ("presentacion_inversores", r"investor present|presentaci[oó]n para inversores|conference call|investor day"),
    ("prospecto", r"prospecto|suplemento de precio|offering memorandum|pricing supplement"),
    ("obligacion_negociable", r"obligaciones negociables"),
    ("informe_calificacion", r"calificaci[oó]n de riesgo|rating action|fix scr|moody|fitch|standard & ?poor"),
    ("reporte_sostenibilidad", r"reporte (de )?(sustentabilidad|sostenibilidad)|reporte asg"),
    ("terminos_y_condiciones", r"t[eé]rminos y condiciones"),
    ("codigo_de_etica", r"c[oó]digo de [eé]tica"),
]

# Real period, read from content. Ordered from most to least specific: a full
# date beats a quarter, a quarter beats a bare year.
# [ES] Periodo real, leido del contenido. De mas a menos especifico: una fecha
# completa le gana a un trimestre, y un trimestre a un ano suelto.
PATRONES_PERIODO = [
    ("fecha_cierre", r"(?:al|cerrado el|finalizado el)\s+(\d{1,2}\s+de\s+\w+\s+de\s+20\d{2})"),
    ("fecha_cierre", r"(?:al|31/|30/)\s?(\d{1,2}[/-]\d{1,2}[/-]20\d{2})"),
    ("trimestre", r"\b([1-4])\s?[QT]\s?(?:20)?(\d{2})\b"),
    ("trimestre", r"\b[QT]\s?([1-4])\s?(?:20)?(\d{2})\b"),
    ("ejercicio", r"ejercicio (?:econ[oó]mico )?(?:finalizado|cerrado)[^\d]{0,20}(20\d{2})"),
    ("anio", r"\b(20[12]\d)\b"),
]

MESES = {
    "enero": "01", "febrero": "02", "marzo": "03", "abril": "04", "mayo": "05",
    "junio": "06", "julio": "07", "agosto": "08", "septiembre": "09",
    "setiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12",
}


def plano(t: str) -> str:
    d = unicodedata.normalize("NFKD", t or "")
    s = "".join(c for c in d if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def leer_pdf(ruta: Path) -> dict:
    """Text of the head and tail pages, page count, and per-page text.

    Reading every page of a 13 MB annual report to decide its type would cost
    minutes for nothing: the identity of a document is stated at the front, and
    the signatures are at the back.

    [ES] Texto de las paginas de cabeza y cola, cantidad de paginas, y texto por
    pagina.

    Leer las 400 paginas de una memoria de 13 MB para decidir su tipo costaria
    minutos para nada: la identidad de un documento se declara adelante, y las
    firmas estan atras.
    """
    import pypdfium2

    try:
        documento = pypdfium2.PdfDocument(str(ruta))
    except Exception as error:
        return {"error": f"{type(error).__name__}", "paginas": None, "por_pagina": {}}

    total = len(documento)
    indices = list(range(min(PAGINAS_A_LEER, total)))
    indices += [i for i in range(max(0, total - PAGINAS_FINALES), total) if i not in indices]

    por_pagina = {}
    for i in indices:
        try:
            por_pagina[i + 1] = documento[i].get_textpage().get_text_range()
        except Exception:
            por_pagina[i + 1] = ""
    documento.close()

    texto = "\n".join(por_pagina.values())
    return {
        "paginas": total,
        "por_pagina": por_pagina,
        "texto": texto,
        "texto_plano": plano(texto),
        "sha256_texto": hashlib.sha256(plano(texto).encode()).hexdigest(),
        "caracteres": len(texto),
    }


UNIDADES = [
    ("miles", r"(?<![a-z])en miles(?![a-z])|(?<![a-z])thousands?(?![a-z])"),
    ("millones", r"(?<![a-z])en millones(?![a-z])|(?<![a-z])millions?(?![a-z])"),
    ("ARS", r"(?<![a-z])ars(?![a-z])|pesos argentinos|\$ ?ars"),
    ("USD", r"(?<![a-z])usd(?![a-z])|dolares|u\$s"),
    ("porcentaje", r"%"),
]


def leer_xlsx(ruta):
    """Read a spreadsheet natively: sheets, cells, units and cell provenance.

    These two files were left uncharacterised because "there is no lightweight
    reader" - which was wrong: `openpyxl` is a dependency of this project and
    reads them directly. Declaring a file unreadable when the reader is already
    installed is not caution, it is a gap in the instrument.

    A spreadsheet is not a PDF and pretending otherwise loses what makes it
    useful: the evidence here is a CELL, with its sheet and coordinate.

    [ES] Lee una planilla nativamente: hojas, celdas, unidades y procedencia de
    celda.

    Estos dos archivos habian quedado sin caracterizar porque "no hay lector
    liviano", que era falso: `openpyxl` es dependencia de este proyecto y los lee
    directamente. Declarar ilegible un archivo cuando el lector ya esta instalado
    no es cautela, es un hueco del instrumento.

    Una planilla no es un PDF y aparentar lo contrario pierde lo que la vuelve
    util: aca la evidencia es una CELDA, con su hoja y su coordenada.
    """
    from openpyxl import load_workbook

    try:
        libro = load_workbook(ruta, read_only=True, data_only=True)
    except Exception as error:
        return {"error": type(error).__name__, "paginas": None, "por_pagina": {}}

    hojas, por_hoja, celdas_texto = [], {}, []
    for nombre in libro.sheetnames:
        hoja = libro[nombre]
        textos, no_vacias, numericas = [], 0, 0
        for fila in hoja.iter_rows(values_only=False):
            for celda in fila:
                if celda.value is None:
                    continue
                no_vacias += 1
                if isinstance(celda.value, (int, float)):
                    numericas += 1
                    continue
                texto = str(celda.value).strip()
                if texto:
                    textos.append(texto)
                    if len(celdas_texto) < 4000:
                        celdas_texto.append(
                            {"hoja": nombre, "coordenada": celda.coordinate,
                             "texto": texto[:200]}
                        )
        hojas.append({
            "hoja": nombre, "celdas_no_vacias": no_vacias,
            "celdas_numericas": numericas, "dimensiones": hoja.calculate_dimension(),
        })
        por_hoja[nombre] = " \n".join(textos)
    libro.close()

    texto = " \n".join(por_hoja.values())
    unidades = [u for u, patron in UNIDADES if re.search(patron, plano(texto))]

    return {
        "paginas": len(hojas),
        "hojas": hojas,
        "por_pagina": por_hoja,
        "texto": texto,
        "texto_plano": plano(texto),
        "sha256_texto": hashlib.sha256(plano(texto).encode()).hexdigest(),
        "caracteres": len(texto),
        "unidades_detectadas": unidades,
        "celdas_muestra": celdas_texto[:12],
    }


def oracion_con(termino: str, por_pagina: dict) -> tuple:
    """The sentence containing a term, and its page. Verbatim, for the reviewer.

    [ES] La oracion que contiene un termino, y su pagina. Textual, para el
    revisor.
    """
    patron = _patron_de_termino(termino)
    for pagina, texto in por_pagina.items():
        m = patron.search(plano(texto))
        if m is None:
            continue
        pos = m.start()
        inicio = max(0, pos - 90)
        fragmento = " ".join(texto[inicio: pos + 150].split())
        return fragmento, pagina
    return "", None


def _patron_de_termino(termino: str) -> re.Pattern:
    """A term matched at lexical boundaries, so `iva` is not found in `derivada`.

    Word boundaries alone are not enough for multi-word terms with accents
    already stripped, so the guard is on alphanumerics either side.

    [ES] Un termino buscado en limites lexicos, para que `iva` no aparezca dentro
    de `derivada`.

    Los limites de palabra por si solos no alcanzan para terminos de varias
    palabras con los acentos ya sacados, asi que la guarda es sobre
    alfanumericos a cada lado.
    """
    return re.compile(r"(?<![a-z0-9])" + re.escape(termino) + r"(?![a-z0-9])")


def proponer_dominios(lectura: dict) -> dict:
    """Per domain: distinct terms, total mentions, pages, and the evidence.

    A domain is proposed only when it clears all three thresholds. One term
    appearing once on one page is an incidental mention, not subject matter: an
    annual report that names ENARGAS in a list of counterparties is not a
    regulatory document.

    Below any threshold the counts are still reported, with the reason it did
    not qualify. A reviewer may well decide that two very specific terms outweigh
    five generic ones, and hiding the numbers would take that decision away.

    [ES] Por dominio: terminos distintos, menciones totales, paginas, y la
    evidencia.

    Un dominio se propone solo si supera los tres umbrales. Un termino que
    aparece una vez en una pagina es una mencion incidental, no materia: una
    memoria que nombra a ENARGAS en una lista de contrapartes no es un documento
    regulatorio.

    Por debajo de cualquier umbral los conteos igual se informan, con el motivo
    por el que no califico. Un revisor bien puede decidir que dos terminos muy
    especificos pesan mas que cinco genericos, y esconder los numeros le quitaria
    esa decision.
    """
    salida = {}
    por_pagina = lectura.get("por_pagina", {})
    planos = {p: plano(t) for p, t in por_pagina.items()}

    for dominio, terminos in TERMINOS.items():
        encontrados, evidencias = [], []
        ocurrencias_totales = 0
        paginas_por_termino = {}

        for termino in terminos:
            patron = _patron_de_termino(termino)
            paginas, cuenta = set(), 0
            for pagina, texto in planos.items():
                hallazgos = patron.findall(texto)
                if hallazgos:
                    paginas.add(pagina)
                    cuenta += len(hallazgos)
            if not cuenta:
                continue
            encontrados.append(termino)
            ocurrencias_totales += cuenta
            paginas_por_termino[termino] = sorted(paginas)
            if len(evidencias) < 3:
                fragmento, pagina = oracion_con(termino, por_pagina)
                if fragmento:
                    evidencias.append({
                        "termino": termino, "pagina": pagina, "cita": fragmento,
                        "ocurrencias": cuenta, "paginas": len(paginas),
                    })

        max_paginas = max((len(v) for v in paginas_por_termino.values()), default=0)
        propuesto = (
            len(encontrados) >= UMBRAL_TERMINOS
            and ocurrencias_totales >= UMBRAL_OCURRENCIAS
            and max_paginas >= UMBRAL_PAGINAS
        )
        salida[dominio] = {
            "terminos_distintos": len(encontrados),
            "ocurrencias_totales": ocurrencias_totales,
            "max_paginas_de_un_termino": max_paginas,
            "terminos": encontrados[:8],
            "propuesto": propuesto,
            "motivo_no_propuesto": None if propuesto else (
                "terminos=%d/%d ocurrencias=%d/%d paginas=%d/%d"
                % (len(encontrados), UMBRAL_TERMINOS,
                   ocurrencias_totales, UMBRAL_OCURRENCIAS,
                   max_paginas, UMBRAL_PAGINAS)
            ),
            "evidencia": evidencias,
        }
    return salida


def proponer_tipo(lectura: dict, nombre: str) -> dict:
    """Document type, with the FILE NAME taking priority over the body.

    `2025 Annual Report.pdf` was typed `estado_financiero`, because an annual
    report contains financial statements and the body was searched first. The
    type then failed to trigger the annual-period rule, and the document ended
    up dated `2T2026` from a stray quarter on its cover.

    A publisher naming a file `Annual Report` is telling us what it is. The body
    only decides when the name says nothing.

    [ES] Tipo documental, con el NOMBRE DE ARCHIVO por encima del cuerpo.

    `2025 Annual Report.pdf` quedaba tipado `estado_financiero`, porque una
    memoria contiene estados financieros y el cuerpo se buscaba primero. El tipo
    entonces no disparaba la regla de periodo anual, y el documento terminaba
    fechado `2T2026` por un trimestre suelto de su tapa.

    Un editor que nombra un archivo `Annual Report` nos esta diciendo que es. El
    cuerpo decide solo cuando el nombre no dice nada.
    """
    del_nombre = plano(nombre)
    for tipo, patron in TIPOS:
        if re.search(patron, del_nombre):
            return {"tipo": tipo, "fuente": "nombre_de_archivo"}

    del_cuerpo = lectura.get("texto_plano", "")[:6000]
    for tipo, patron in TIPOS:
        if re.search(patron, del_cuerpo):
            return {"tipo": tipo, "fuente": "contenido"}
    return {"tipo": "no_determinado", "fuente": None}


def _normalizar_periodo(clase, m) -> str:
    if clase == "trimestre":
        t, a = m.group(1), m.group(2)
        return f"{t}T20{a}" if len(a) == 2 else f"{t}T{a}"
    valor = m.group(1)
    # `re.I` is the whole fix. Without it, `31 DE MARZO DE 2026` did not match
    # the lowercase `de` and fell through unnormalised, so fourteen closing
    # dates stayed as free text and two documents of the same period compared
    # as different.
    # [ES] `re.I` es toda la correccion. Sin eso, `31 DE MARZO DE 2026` no
    # matcheaba el `de` en minuscula y salia sin normalizar, asi que catorce
    # fechas de cierre quedaban como texto libre y dos documentos del mismo
    # periodo comparaban distinto.
    m2 = re.match(r"(\d{1,2})\s+de\s+(\w+)\s+de\s+(20\d{2})", valor, re.I)
    if m2:
        mes = MESES.get(plano(m2.group(2)))
        if mes:
            return f"{m2.group(3)}-{mes}-{int(m2.group(1)):02d}"
    m3 = re.match(r"(\d{1,2})[/-](\d{1,2})[/-](20\d{2})", valor)
    if m3:
        return f"{m3.group(3)}-{int(m3.group(2)):02d}-{int(m3.group(1)):02d}"
    return valor


def proponer_periodo(lectura: dict, nombre: str, texto_enlace: str, url: str,
                     tipo_propuesto: str = "") -> dict:
    """Period by priority of source, and NEVER from the URL directory alone.

    `/uploads/2026/07/EEFF-31-12-25.pdf` is a 2025 statement published in 2026.
    Taking the directory year would date it wrong by a year, quietly, on every
    document a site publishes late.

    The URL year is still computed - and reported as a CONTRAST, so a reviewer
    can see when publication and period diverge.

    [ES] Periodo por prioridad de fuente, y NUNCA del directorio de la URL solo.

    `/uploads/2026/07/EEFF-31-12-25.pdf` es un estado de 2025 publicado en 2026.
    Tomar el ano del directorio lo fecharia mal por un ano, en silencio, en todo
    documento que un sitio publique con retraso.

    El ano de la URL igual se calcula, y se informa como CONTRASTE, para que un
    revisor vea cuando publicacion y periodo divergen.
    """
    def buscar(texto):
        for clase, patron in PATRONES_PERIODO:
            m = re.search(patron, texto, re.I)
            if m:
                return {"valor": _normalizar_periodo(clase, m), "clase": clase}
        return None

    candidatos = []
    desde_contenido = buscar(lectura.get("texto", "")[:20000])
    if desde_contenido:
        candidatos.append(dict(desde_contenido, fuente="contenido_del_pdf"))
    desde_nombre = buscar(nombre)
    if desde_nombre:
        candidatos.append(dict(desde_nombre, fuente="nombre_de_archivo"))
    desde_enlace = buscar(texto_enlace or "")
    if desde_enlace:
        candidatos.append(dict(desde_enlace, fuente="texto_del_enlace"))

    anio_url = None
    m = re.search(r"/(20[12]\d)/", url or "")
    if m:
        anio_url = m.group(1)

    # Not the first candidate found, but the most reliable CLASS of evidence.
    # `2025 Annual Report` matched a quarter somewhere in its head text and was
    # dated `2T2026`; an explicit "ejercicio finalizado" or a closing date is
    # worth more than a stray quarter, whatever order they appear in.
    # [ES] No el primer candidato encontrado, sino la CLASE de evidencia mas
    # confiable. `2025 Annual Report` matcheo un trimestre en algun lugar de su
    # texto de cabecera y quedo fechado `2T2026`; un "ejercicio finalizado"
    # explicito o una fecha de cierre valen mas que un trimestre suelto,
    # aparezcan en el orden que aparezcan.
    PESO_CLASE = {"fecha_cierre": 4, "ejercicio": 3, "trimestre": 2, "anio": 1}
    PESO_FUENTE = {"contenido_del_pdf": 3, "nombre_de_archivo": 2, "texto_del_enlace": 1}

    utiles = list(candidatos)
    # An annual document has no quarter. Keeping one would date a memoria by the
    # first quarter its cover happens to mention.
    # [ES] Un documento anual no tiene trimestre. Conservarlo fecharia una
    # memoria por el primer trimestre que su tapa mencione.
    if tipo_propuesto in ("memoria_anual", "reporte_sostenibilidad"):
        sin_trimestre = [c for c in utiles if c["clase"] != "trimestre"]
        if sin_trimestre:
            utiles = sin_trimestre

    elegido = max(
        utiles,
        key=lambda c: (PESO_CLASE.get(c["clase"], 0), PESO_FUENTE.get(c["fuente"], 0)),
    ) if utiles else None
    return {
        "periodo_propuesto": elegido["valor"] if elegido else None,
        "fuente": elegido["fuente"] if elegido else None,
        "clase": elegido["clase"] if elegido else None,
        "candidatos": candidatos,
        "anio_en_url": anio_url,
        "discrepa_con_url": bool(
            elegido and anio_url and anio_url not in str(elegido["valor"])
        ),
        "fuentes_discrepan": len({c["valor"] for c in candidatos}) > 1,
        "confianza": (
            "alta" if elegido and elegido["clase"] in ("fecha_cierre", "ejercicio")
            and elegido["fuente"] == "contenido_del_pdf"
            else "media" if elegido else "sin_periodo"
        ),
    }


def detectar_no_pertinente(lectura: dict, nombre: str) -> list:
    contexto = plano(nombre) + " " + lectura.get("texto_plano", "")[:4000]
    return [etiqueta for etiqueta, patron in PATRONES_NO_PERTINENTES
            if re.search(patron, contexto)]


# Words that never start a company name. A capture beginning with one of these
# is a sentence that happens to mention a company, not the company.
# [ES] Palabras que nunca inician una razon social. Una captura que empieza con
# alguna de estas es una oracion que menciona una empresa, no la empresa.
ARRANQUES_DE_ORACION = (
    # Connectors and pronouns.
    # [ES] Conectores y pronombres.
    "to", "on", "the", "of", "and", "for", "in", "at", "by", "from", "with",
    "we", "our", "its", "this", "these", "as",
    "por", "de", "del", "la", "el", "los", "las", "una", "un", "en", "a",
    # Addressees. `To the shareholders of Pampa Energia S.A` survives the
    # connector trim as `shareholders of Pampa Energia S.A`, which is still a
    # clause. No company is named after the people a letter is addressed to.
    #
    # Trimming at `of` instead would look tidier and would be wrong: `GAS
    # TRANSPORTER OF THE SOUTH INC` is a real name in this corpus, and cutting
    # there would leave `THE SOUTH INC`.
    #
    # [ES] Destinatarios. `To the shareholders of Pampa Energia S.A` sobrevive al
    # recorte de conectores como `shareholders of Pampa Energia S.A`, que sigue
    # siendo una oracion. Ninguna empresa se llama como las personas a las que se
    # dirige una carta.
    #
    # Recortar en `of` se veria mas prolijo y estaria mal: `GAS TRANSPORTER OF
    # THE SOUTH INC` es un nombre real de este corpus, y cortar ahi dejaria `THE
    # SOUTH INC`.
    "shareholders", "stockholders", "directors", "board", "members",
    "senores", "accionistas", "senor", "sres", "directorio", "asamblea",
    "informe", "memoria", "estados", "notas", "anexo",
)

# A company name has at most this many words. `To the shareholders of Pampa
# Energia S.A` has seven and is a clause; `Transportadora de Gas del Sur S.A.`
# has six and is a name. The cut is at six, and anything longer is trimmed from
# the left rather than accepted whole.
# [ES] Una razon social tiene a lo sumo esta cantidad de palabras. `To the
# shareholders of Pampa Energia S.A` tiene siete y es una oracion;
# `Transportadora de Gas del Sur S.A.` tiene seis y es un nombre. El corte esta
# en seis, y lo mas largo se recorta desde la izquierda en lugar de aceptarse
# entero.
MAXIMO_PALABRAS_RAZON_SOCIAL = 6


def _limpiar_razon_social(crudo: str):
    """Turn a capture into a company name, or into nothing.

    Three ways a capture fails to be a name, all of them seen on the real
    corpus:

      `To the shareholders of Pampa Energia S.A`  a clause containing a name
      `On April 21, 2026, Fertil Pampa S.A.U`     a sentence with a date in it
      `MMESA Compania Administradora del ...`     longer than any real name

    A clause is trimmed from the left, keeping the tail, and only accepted if
    what remains still looks like a name. Anything with digits is refused
    outright: no company name contains a year.

    [ES] Convierte una captura en una razon social, o en nada.

    Tres formas en que una captura no llega a ser un nombre, las tres vistas en
    el corpus real. Una oracion se recorta desde la izquierda conservando la
    cola, y solo se acepta si lo que queda sigue pareciendo un nombre. Cualquier
    cosa con digitos se rechaza de plano: ninguna razon social lleva un ano.
    """
    texto = " ".join((crudo or "").split())
    if not texto:
        return None
    # A year, a day number or any digit run: this is a sentence, not a name.
    # [ES] Un ano, un numero de dia o cualquier corrida de digitos: esto es una
    # oracion, no un nombre.
    if re.search(r"\d", texto):
        partes = texto.split()
        while partes and re.search(r"\d", partes[0]):
            partes.pop(0)
        # Drop everything up to and including the last token with a digit.
        # [ES] Descartar todo hasta el ultimo token con digitos, inclusive.
        indices = [i for i, t in enumerate(partes) if re.search(r"\d", t)]
        if indices:
            partes = partes[indices[-1] + 1:]
        texto = " ".join(partes)
        if not texto:
            return None

    partes = texto.split()
    # Trim leading sentence words until the capture starts like a name.
    # [ES] Recortar palabras iniciales de oracion hasta que la captura empiece
    # como un nombre.
    while partes and plano(partes[0]).strip(",.") in ARRANQUES_DE_ORACION:
        partes.pop(0)
    if not partes:
        return None
    if len(partes) > MAXIMO_PALABRAS_RAZON_SOCIAL:
        partes = partes[-MAXIMO_PALABRAS_RAZON_SOCIAL:]
        while partes and plano(partes[0]).strip(",.") in ARRANQUES_DE_ORACION:
            partes.pop(0)

    texto = " ".join(partes).strip(" ,.")
    if len(plano(texto)) < 6:
        return None
    # After trimming, the capture must still carry the legal-form marker that
    # made it a candidate in the first place.
    # [ES] Despues de recortar, la captura tiene que seguir llevando la marca de
    # forma juridica que la volvio candidata en primer lugar.
    if not re.search(r"s\.?a\.?", plano(texto)):
        return None
    return texto


def proponer_entidad(lectura: dict, sugerida: str) -> dict:
    """Entity from provenance when there is one, confirmed against the text.

    A file taken from an issuer's own investor page is strong provenance. A file
    in quarantine has none, and the entity has to be read from the document - or
    left absent, which is a valid answer and a common one.

    [ES] Entidad desde la procedencia cuando la hay, contrastada con el texto.

    Un archivo tomado de la pagina de inversores de la propia emisora es
    procedencia fuerte. Un archivo en cuarentena no tiene ninguna, y la entidad
    hay que leerla del documento, o dejarla ausente, que es una respuesta valida
    y frecuente.
    """
    texto = lectura.get("texto", "") or ""
    crudas = re.findall(
        r"([A-ZÁÉÍÓÚÑ][\w&.,\- ]{3,60}?\s+S\.?A\.?(?:\.?U\.?)?)\b", texto[:6000]
    )

    limpias, vistas = [], set()
    for cruda in crudas:
        nombre = _limpiar_razon_social(cruda)
        if not nombre:
            continue
        clave = plano(nombre)
        if clave in vistas:
            continue
        vistas.add(clave)
        limpias.append(nombre)

    confirmada = None
    if sugerida:
        primera = plano(sugerida).split()[0] if plano(sugerida).split() else ""
        for nombre in limpias:
            if primera and primera in plano(nombre):
                confirmada = nombre
                break

    # The document names its own entity more than it names a counterparty or a
    # subsidiary. Taking the first capture instead made `Q1 26 Earnings
    # release.pdf` belong to `Fertil Pampa S.A.U` - a subsidiary mentioned once
    # in the opening sentence - rather than to the issuer.
    # [ES] Un documento nombra a su propia entidad mas que a una contraparte o a
    # una subsidiaria. Tomar la primera captura hacia que `Q1 26 Earnings
    # release.pdf` perteneciera a `Fertil Pampa S.A.U` -una subsidiaria
    # mencionada una vez en la oracion inicial- en lugar de a la emisora.
    plano_texto = plano(texto)
    frecuencias = {n: plano_texto.count(plano(n)) for n in limpias}
    ordenadas = sorted(limpias, key=lambda n: (-frecuencias[n], limpias.index(n)))

    return {
        "entidad_propuesta": sugerida or (ordenadas[0] if ordenadas else None),
        "frecuencia_en_texto": frecuencias.get(ordenadas[0]) if ordenadas else 0,
        "origen": "procedencia_ir" if sugerida else (
            "texto_del_documento" if limpias else "ausente"
        ),
        "confirmada_en_texto": bool(confirmada),
        "razones_sociales_detectadas": ordenadas[:5],
        "confianza_entidad": (
            "alta" if sugerida and confirmada else
            "media" if sugerida or (limpias and len(limpias) == 1) else
            "baja" if limpias else "sin_entidad"
        ),
    }


def clave_documental(registro: dict) -> str:
    """Documentary identity: entity + type + period + title + pages.

    Two renderings of the same statement have different bytes and the same
    identity. SHA-256 alone would call them two documents.

    [ES] Identidad documental: entidad + tipo + periodo + titulo + paginas.

    Dos renderizados del mismo estado tienen distintos bytes y la misma
    identidad. El SHA-256 solo los llamaria dos documentos.
    """
    return "|".join([
        plano(registro.get("entidad_propuesta") or "?"),
        registro.get("tipo_propuesto") or "?",
        str(registro.get("periodo_propuesto") or "?"),
        plano(registro.get("titulo") or "")[:60],
        str(registro.get("paginas") or "?"),
    ])


def caracterizar(ruta: Path, zona: str, procedencia: dict) -> dict:
    # A spreadsheet gets read as a spreadsheet. Its evidence is a cell with a
    # sheet and a coordinate, not a page.
    # [ES] Una planilla se lee como planilla. Su evidencia es una celda con hoja
    # y coordenada, no una pagina.
    es_planilla = ruta.suffix.lower() in (".xlsx", ".xlsm")
    lectura = leer_xlsx(ruta) if es_planilla else leer_pdf(ruta)
    nombre = ruta.name
    tipo = proponer_tipo(lectura, nombre)
    entidad = proponer_entidad(lectura, procedencia.get("emisor_nombre"))
    periodo = proponer_periodo(
        lectura, nombre, procedencia.get("titulo_propuesto", ""),
        procedencia.get("url", ""), tipo["tipo"],
    )
    dominios = proponer_dominios(lectura)
    propuestos = [d for d, v in dominios.items() if v["propuesto"]]

    # A `Memoria y EEFF` that devotes a chapter to sustainability is not a
    # sustainability report, and flagging it for exclusion would drop one of the
    # richest documents in the corpus. The flag only stands when the document
    # ALSO carries no regulated-domain material, or when its own type is the
    # non-pertinent one.
    # [ES] Una `Memoria y EEFF` que le dedica un capitulo a la sostenibilidad no
    # es un reporte de sostenibilidad, y marcarla para exclusion tiraria uno de
    # los documentos mas ricos del corpus. La marca solo se sostiene cuando el
    # documento ADEMAS no trae materia de dominio regulado, o cuando su propio
    # tipo es el no pertinente.
    marcas = detectar_no_pertinente(lectura, nombre)
    TIPOS_NO_PERTINENTES = {
        "terminos_y_condiciones", "codigo_de_etica", "reporte_sostenibilidad",
    }
    if marcas and propuestos and tipo["tipo"] not in TIPOS_NO_PERTINENTES:
        marcas = []
    no_pertinente = marcas

    registro = {
        "zona": zona,
        "archivo": nombre,
        "ruta": str(ruta.relative_to(PROJECT_ROOT)).replace("\\", "/"),
        "bytes": ruta.stat().st_size,
        "sha256": hashlib.sha256(ruta.read_bytes()).hexdigest(),
        "formato": "xlsx" if es_planilla else "pdf",
        "paginas": lectura.get("paginas"),
        "hojas": lectura.get("hojas"),
        "unidades_detectadas": lectura.get("unidades_detectadas", []),
        "celdas_muestra": lectura.get("celdas_muestra", []),
        "sha256_texto": lectura.get("sha256_texto"),
        "caracteres_leidos": lectura.get("caracteres"),
        "error_lectura": lectura.get("error"),
        "titulo": procedencia.get("titulo_propuesto") or nombre,
        "url_origen": procedencia.get("url"),
        "fecha_acceso": procedencia.get("fecha_acceso"),
        "sigla": procedencia.get("sigla"),
        "segmento": procedencia.get("segmento"),
        **entidad,
        "tipo_propuesto": tipo["tipo"],
        "tipo_fuente": tipo["fuente"],
        **periodo,
        "dominios": dominios,
        "dominios_propuestos": propuestos,
        "marcas_no_pertinente": no_pertinente,
        "recomendacion": (
            "excluir_candidato" if no_pertinente else
            "revisar" if not propuestos else
            "incluir_candidato"
        ),
        "estado": "pendiente_revision",
    }
    registro["clave_documental"] = clave_documental(registro)
    return registro


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument(
        "--salida", type=Path,
        default=DATA_DIR / "catalog" / "candidates" / "caracterizacion_fase3v2.jsonl",
    )
    args = parser.parse_args()

    adquisicion = json.loads(ADQUISICION.read_text(encoding="utf-8"))
    por_ruta = {
        r["ruta"]: r for r in adquisicion["registros"] if r.get("resultado") == "descargado"
    }

    registros = []
    print("leyendo empresariales descargados ...", flush=True)
    for ruta in sorted(CANDIDATOS.glob("*.pdf")):
        rel = str(ruta.relative_to(PROJECT_ROOT)).replace("\\", "/")
        registros.append(caracterizar(ruta, "incoming_candidates", por_ruta.get(rel, {})))
        unidad = "hojas" if registros[-1].get("formato") == "xlsx" else "pág."
        print(f"  {ruta.name[:60]:62} {registros[-1]['paginas']} {unidad}", flush=True)

    print("leyendo cuarentena ...", flush=True)
    for ruta in sorted(CUARENTENA.iterdir()):
        if not ruta.is_file() or ruta.suffix.lower() not in (".pdf", ".xlsx", ".xlsm"):
            continue
        registros.append(caracterizar(ruta, "cuarentena", {}))
        unidad = "hojas" if registros[-1].get("formato") == "xlsx" else "pág."
        print(f"  {ruta.name[:60]:62} {registros[-1]['paginas']} {unidad}", flush=True)

    # Documentary duplication, beyond bytes.
    # [ES] Duplicacion documental, mas alla de los bytes.
    por_clave = collections.defaultdict(list)
    por_texto = collections.defaultdict(list)
    for r in registros:
        por_clave[r["clave_documental"]].append(r["archivo"])
        if r.get("sha256_texto"):
            por_texto[r["sha256_texto"]].append(r["archivo"])

    duplicados = {
        "por_clave_documental": {k: v for k, v in por_clave.items() if len(v) > 1},
        "por_texto": {k: v for k, v in por_texto.items() if len(v) > 1},
    }

    args.salida.parent.mkdir(parents=True, exist_ok=True)
    with args.salida.open("w", encoding="utf-8") as f:
        for r in registros:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    resumen = collections.Counter(r["recomendacion"] for r in registros)
    print()
    print(f"documentos caracterizados: {len(registros)}")
    for k, v in resumen.most_common():
        print(f"  {k:22} {v}")
    print(f"  duplicados por clave documental: {len(duplicados['por_clave_documental'])}")
    print(f"  duplicados por texto:            {len(duplicados['por_texto'])}")
    print(f"\njsonl  {args.salida}")


if __name__ == "__main__":
    main()
