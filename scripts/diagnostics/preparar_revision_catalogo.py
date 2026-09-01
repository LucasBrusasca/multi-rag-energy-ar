"""Build the human review instrument for the 24 canonical documents.

WHY THIS EXISTS. All 24 catalogue records carry `estado_inclusion =
pendiente_revision`, so their `dominios_documentales`, `tipo_documento` and
`emisor_id` are descriptive proposals, not confirmed human truth. Every number
that leans on them - RQ0 included - inherits that. This sheet is the instrument
for turning the proposals into decisions, one document at a time.

WHAT "VERIFIABLE EVIDENCE" MEANS HERE, AND WHAT IT DOES NOT. The evidence
columns point at WHERE TO LOOK. They do not argue that a label is right.

- `evidencia_extracto` is a VERBATIM excerpt of one chunk, with its `chunk_uid`
  and page, so the claim can be reopened in the source document.
- `evidencia_por_dominio` reports, for each currently assigned domain, whether
  that word appears LITERALLY in the document and where. It is a string match on
  the token itself, not an invented lexicon and not a judgement: "the word
  `financiero` appears on page 4" is a fact; "therefore the document is
  financial" is Lucas's call, not this script's. A `sin coincidencia literal` is
  informative in its own right - a document labelled with a domain whose word
  never appears is worth a second look.

NOTHING IS DECIDED HERE. The script does not change a single label, does not
touch PostgreSQL and does not write to the catalogue. The decision column, the
four correction columns and `observaciones` come out EMPTY, and filling them is
the whole point.

ONE COLUMN PER CORRECTABLE FIELD. A single free-text correction box forces the
reviewer to retype a whole record to change one thing, and forces whoever
applies it later to read prose to find out what changed. With one column per
field the diff is explicit: an empty cell was not corrected and keeps its
current value.

INGESTION DEFECTS ARE REGISTERED, NOT REPAIRED. Replacement characters (U+FFFD)
found in the corpus are listed on their own sheet with document, chunk and page.
They are a defect of ingestion, not a labelling question, and silently fixing
them here would hide a corpus problem inside a metadata review.

WHY THE AUTOMATIC CLASSIFICATION IS ON A SEPARATE SHEET. `chunks.silo` holds
what the classifier decided. Putting it beside the label under review would
anchor the reviewer toward agreeing with the machine, which is exactly the
failure mode that makes a human reference worthless. It is available in
`contexto_automatico` for afterwards.

[ES] Arma el instrumento de revision humana para los 24 documentos canonicos.

POR QUE EXISTE. Los 24 registros del catalogo tienen `estado_inclusion =
pendiente_revision`, asi que sus `dominios_documentales`, `tipo_documento` y
`emisor_id` son propuestas descriptivas, no verdad humana confirmada. Todo numero
que se apoye en ellos -RQ0 incluido- hereda eso. Esta planilla es el instrumento
para convertir las propuestas en decisiones, documento por documento.

QUE SIGNIFICA "EVIDENCIA VERIFICABLE" ACA, Y QUE NO. Las columnas de evidencia
senalan DONDE MIRAR. No argumentan que una etiqueta sea correcta.

- `evidencia_extracto` es un extracto TEXTUAL de un chunk, con su `chunk_uid` y
  su pagina, para poder reabrir la afirmacion en el documento fuente.
- `evidencia_por_dominio` informa, para cada dominio asignado hoy, si esa palabra
  aparece LITERALMENTE en el documento y donde. Es una busqueda de la palabra
  misma, no un lexico inventado ni un juicio: "la palabra `financiero` aparece en
  la pagina 4" es un hecho; "por lo tanto el documento es financiero" lo decide
  Lucas, no este script. Un `sin coincidencia literal` es informativo por si
  solo: un documento etiquetado con un dominio cuya palabra nunca aparece merece
  una segunda mirada.

ACA NO SE DECIDE NADA. El script no cambia ninguna etiqueta, no toca PostgreSQL
y no escribe en el catalogo. La columna de decision, las cuatro de correccion y
`observaciones` salen VACIAS, y llenarlas es todo el punto.

UNA COLUMNA POR CAMPO CORREGIBLE. Una unica caja de texto libre obliga al revisor
a reescribir un registro entero para cambiar una sola cosa, y obliga a quien
despues lo aplique a leer prosa para averiguar que cambio. Con una columna por
campo el diff es explicito: una celda vacia no se corrigio y conserva su valor
actual.

LOS DEFECTOS DE INGESTA SE REGISTRAN, NO SE REPARAN. Los caracteres de reemplazo
(U+FFFD) encontrados en el corpus se listan en su propia hoja, con documento,
chunk y pagina. Son un defecto de ingesta, no una pregunta de etiquetado, y
arreglarlos en silencio aca escondería un problema del corpus dentro de una
revision de metadatos.

POR QUE LA CLASIFICACION AUTOMATICA VA EN UNA HOJA APARTE. `chunks.silo` tiene
lo que decidio el clasificador. Ponerlo al lado de la etiqueta bajo revision
ancliaria al revisor hacia coincidir con la maquina, que es justamente el modo de
falla que vuelve inservible una referencia humana. Queda en
`contexto_automatico` para despues.
"""

import argparse
import collections
import csv
import hashlib
import json
import re
import unicodedata
from pathlib import Path

from multirag.config import SILOS
from multirag.db import conectar
from multirag.paths import DATA_DIR, EXPERIMENTS_DIR


RECETA_VERSION = "revision-catalogo-v2"

CATALOGO = DATA_DIR / "catalog" / "metadatos_curados.csv"
INVENTARIO = DATA_DIR / "catalog" / "inventario_objetivo.jsonl"

SEPARADOR_DOMINIOS = "|"
LARGO_EXTRACTO = 320

COLUMNAS = (
    "document_id",
    "fuente",
    "titulo_oficial",
    "archivo_referencia",
    "emisor_id",
    "emisor_nombre",
    "tipo_documento",
    "dominios_documentales",
    "jurisdiccion",
    "fecha_documento",
    "chunks",
    "paginas",
    "evidencia_extracto",
    "evidencia_por_dominio",
    "url_origen",
    "estado_inclusion",
    # Empty on purpose. Filling them is the review.
    #
    # One column per correctable field, instead of one free-text box for all of
    # them. A single `correccion_humana` forces the reviewer to re-type the
    # whole record to change one thing, and forces whoever applies it later to
    # parse prose to find out what changed. Separate columns make the diff
    # explicit: what is empty was not corrected, and it keeps its current value.
    #
    # [ES] Vacias a proposito.
    #
    # Una columna por campo corregible, en lugar de una caja de texto libre para
    # todos. Un unico `correccion_humana` obliga al revisor a reescribir el
    # registro entero para cambiar una sola cosa, y obliga a quien despues lo
    # aplique a interpretar prosa para averiguar que cambio. Columnas separadas
    # vuelven explicito el diff: lo que esta vacio no se corrigio, y conserva su
    # valor actual.
    "decision_humana",
    "emisor_id_corregido",
    "emisor_nombre_corregido",
    "tipo_documento_corregido",
    "dominios_documentales_corregidos",
    "observaciones",
)

CAMPOS_CORREGIBLES = (
    "emisor_id_corregido",
    "emisor_nombre_corregido",
    "tipo_documento_corregido",
    "dominios_documentales_corregidos",
)

COLUMNAS_A_COMPLETAR = ("decision_humana",) + CAMPOS_CORREGIBLES + ("observaciones",)

OPCIONES_DECISION = ("confirmar", "corregir", "dudoso", "excluir_del_corpus")


def huella_de_archivo(ruta: Path) -> str:
    h = hashlib.sha256()
    with ruta.open("rb") as f:
        for bloque in iter(lambda: f.read(1 << 20), b""):
            h.update(bloque)
    return f"sha256:{h.hexdigest()}"


def _plano(texto: str) -> str:
    """Lowercase without accents, for the literal word search only.

    [ES] Minusculas sin acentos, solo para la busqueda literal de la palabra.
    """
    d = unicodedata.normalize("NFKD", texto or "")
    return "".join(c for c in d if not unicodedata.combining(c)).lower()


def leer_inventario() -> set:
    with INVENTARIO.open(encoding="utf-8") as f:
        return {json.loads(l)["artifact_id"] for l in f if l.strip()}


def leer_catalogo() -> list:
    with CATALOGO.open(encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def leer_chunks(artifacts_objetivo: set) -> dict:
    """All chunks of the 24 canonical documents, in a READ ONLY transaction.

    [ES] Todos los chunks de los 24 documentos canonicos, en una transaccion
    READ ONLY.
    """
    conexion = conectar()
    try:
        conexion.set_session(readonly=True)
        cur = conexion.cursor()
        cur.execute(
            """
            select document_id, artifact_id, chunk_uid, titulo, contenido,
                   paginas, silo
            from chunks
            order by document_id, chunk_uid
            """
        )
        filas = cur.fetchall()
    finally:
        conexion.close()

    por_documento = collections.defaultdict(list)
    for doc, artifact, uid, titulo, contenido, paginas, silo in filas:
        if artifact not in artifacts_objetivo:
            continue
        por_documento[doc].append(
            {
                "chunk_uid": uid,
                "titulo": titulo,
                "contenido": contenido or "",
                "paginas": list(paginas or []),
                "silo": silo,
            }
        )
    return dict(por_documento)


def _pagina(chunk) -> str:
    return str(chunk["paginas"][0]) if chunk["paginas"] else "s/p"


def extracto_verificable(chunks) -> str:
    """A verbatim excerpt with its chunk_uid and page, so it can be reopened.

    The first substantive chunk by page order. Nothing is paraphrased: what is
    quoted is what the document says, cut at a fixed length.

    [ES] Un extracto textual con su chunk_uid y pagina, para poder reabrirlo.

    El primer chunk sustantivo por orden de pagina. No se parafrasea nada: lo
    citado es lo que dice el documento, cortado a un largo fijo.
    """
    if not chunks:
        return "(sin chunks)"
    ordenados = sorted(
        chunks, key=lambda c: (c["paginas"][0] if c["paginas"] else 10**6, c["chunk_uid"])
    )
    elegido = next(
        (c for c in ordenados if len(c["contenido"].strip()) >= 120), ordenados[0]
    )
    texto = " ".join(elegido["contenido"].split())[:LARGO_EXTRACTO]
    titulo = (elegido["titulo"] or "").strip()
    cabeza = f"[{elegido['chunk_uid'][:12]}… p.{_pagina(elegido)}]"
    return f"{cabeza} {titulo + ' — ' if titulo else ''}«{texto}…»"


def evidencia_por_dominio(dominios: str, chunks) -> str:
    """For each assigned domain: does that WORD appear literally, and where?

    A plain string search for the token itself. It is deliberately not a
    domain lexicon: inventing one would smuggle this script's opinion into an
    instrument whose only job is to show the reviewer where to look.

    `sin coincidencia literal` is a finding too - a document labelled with a
    domain whose word never appears deserves a second look, in either direction.

    [ES] Para cada dominio asignado: aparece esa PALABRA literalmente, y donde?

    Una busqueda de texto de la palabra misma. A proposito no es un lexico de
    dominio: inventar uno colaria la opinion de este script en un instrumento
    cuyo unico trabajo es mostrarle al revisor donde mirar.

    `sin coincidencia literal` tambien es un hallazgo: un documento etiquetado
    con un dominio cuya palabra nunca aparece merece una segunda mirada, en
    cualquiera de las dos direcciones.
    """
    tokens = [t.strip() for t in (dominios or "").split(SEPARADOR_DOMINIOS) if t.strip()]
    if not tokens:
        return "(sin dominios asignados)"

    lineas = []
    for token in tokens:
        patron = re.compile(rf"\b{re.escape(_plano(token))}", re.IGNORECASE)
        golpes = []
        for c in chunks:
            if patron.search(_plano(c["contenido"])) or patron.search(_plano(c["titulo"] or "")):
                golpes.append(c)
        if not golpes:
            lineas.append(f"{token}: sin coincidencia literal en el documento")
            continue
        primero = sorted(
            golpes, key=lambda c: (c["paginas"][0] if c["paginas"] else 10**6, c["chunk_uid"])
        )[0]
        lineas.append(
            f"{token}: {len(golpes)} chunk(s); p.ej. p.{_pagina(primero)} "
            f"[{primero['chunk_uid'][:12]}…]"
        )
    return "\n".join(lineas)


def armar_filas(catalogo, chunks_por_doc) -> list:
    filas = []
    for registro in sorted(catalogo, key=lambda r: r["document_id"]):
        doc = registro["document_id"]
        chunks = chunks_por_doc.get(doc, [])
        paginas = sorted({p for c in chunks for p in c["paginas"]})
        filas.append(
            {
                "document_id": doc,
                "fuente": registro.get("fuente", ""),
                "titulo_oficial": registro.get("titulo_oficial", ""),
                "archivo_referencia": registro.get("archivo_referencia", ""),
                "emisor_id": registro.get("emisor_id", ""),
                "emisor_nombre": registro.get("emisor_nombre", ""),
                "tipo_documento": registro.get("tipo_documento", ""),
                "dominios_documentales": registro.get("dominios_documentales", ""),
                "jurisdiccion": registro.get("jurisdiccion", ""),
                "fecha_documento": registro.get("fecha_documento", ""),
                "chunks": len(chunks),
                "paginas": f"{min(paginas)}–{max(paginas)}" if paginas else "s/p",
                "evidencia_extracto": extracto_verificable(chunks),
                "evidencia_por_dominio": evidencia_por_dominio(
                    registro.get("dominios_documentales", ""), chunks
                ),
                "url_origen": registro.get("url_origen", ""),
                "estado_inclusion": registro.get("estado_inclusion", ""),
                "decision_humana": "",
                "emisor_id_corregido": "",
                "emisor_nombre_corregido": "",
                "tipo_documento_corregido": "",
                "dominios_documentales_corregidos": "",
                "observaciones": "",
            }
        )
    return filas


REEMPLAZO = "�"


def defectos_de_ingesta(chunks_por_doc) -> list:
    """Every U+FFFD in the corpus: a decoding loss, registered and not repaired.

    A replacement character means a byte could not be decoded when the document
    was ingested. Rewriting it here would paper over a corpus defect from inside
    a metadata review, and the next re-ingestion would bring it straight back.
    It is listed instead, with enough coordinates to reopen it.

    [ES] Cada U+FFFD del corpus: una perdida de decodificacion, registrada y no
    reparada.

    Un caracter de reemplazo significa que un byte no se pudo decodificar al
    ingerir el documento. Reescribirlo aca taparia un defecto del corpus desde
    adentro de una revision de metadatos, y la proxima reingesta lo traeria de
    vuelta igual. En cambio se lista, con coordenadas suficientes para reabrirlo.
    """
    hallazgos = []
    for doc in sorted(chunks_por_doc):
        for chunk in chunks_por_doc[doc]:
            for campo in ("titulo", "contenido"):
                texto = chunk.get(campo) or ""
                if REEMPLAZO not in texto:
                    continue
                i = texto.index(REEMPLAZO)
                hallazgos.append(
                    {
                        "document_id": doc,
                        "chunk_uid": chunk["chunk_uid"],
                        "pagina": _pagina(chunk),
                        "campo": campo,
                        "ocurrencias": texto.count(REEMPLAZO),
                        "contexto": " ".join(
                            texto[max(0, i - 45): i + 45].split()
                        ),
                    }
                )
    return hallazgos


def contexto_automatico(catalogo, chunks_por_doc) -> list:
    """What the classifier decided, kept OFF the review sheet to avoid anchoring.

    [ES] Lo que decidio el clasificador, FUERA de la hoja de revision para no
    anclar al revisor.
    """
    filas = []
    for registro in sorted(catalogo, key=lambda r: r["document_id"]):
        doc = registro["document_id"]
        conteo = collections.Counter(
            c["silo"] for c in chunks_por_doc.get(doc, []) if c["silo"]
        )
        total = sum(conteo.values())
        filas.append(
            {
                "document_id": doc,
                "fuente": registro.get("fuente", ""),
                "dominios_documentales_catalogo": registro.get("dominios_documentales", ""),
                "silos_de_chunks_persistidos": ", ".join(
                    f"{s} {n} ({n / total * 100:.0f} %)"
                    for s, n in conteo.most_common()
                ) if total else "(sin silo)",
            }
        )
    return filas


# --------------------------------------------------------------------------


def escribir_csv(ruta: Path, filas) -> None:
    """UTF-8 with BOM: Excel on Windows opens it with the accents intact.

    [ES] UTF-8 con BOM: Excel en Windows lo abre con los acentos correctos.
    """
    ruta.parent.mkdir(parents=True, exist_ok=True)
    with ruta.open("w", encoding="utf-8-sig", newline="") as f:
        escritor = csv.DictWriter(f, fieldnames=COLUMNAS, lineterminator="\n")
        escritor.writeheader()
        for fila in filas:
            escritor.writerow(fila)


def escribir_xlsx(ruta: Path, filas, contexto, defectos, manifest) -> None:
    """The comfortable version: frozen header, wrapped text, a dropdown.

    [ES] La version comoda: encabezado fijo, texto con ajuste, un desplegable.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    libro = Workbook()

    # --- instructions first, because they change how the sheet is read ---
    # [ES] Instrucciones primero, porque cambian como se lee la planilla.
    hoja = libro.active
    hoja.title = "instrucciones"
    instrucciones = [
        ("Revisión humana del catálogo — 24 documentos canónicos", True),
        ("", False),
        ("Qué se revisa", True),
        ("Los 24 registros tienen estado_inclusion = pendiente_revision. Sus", False),
        ("dominios_documentales, tipo_documento y emisor_id son PROPUESTAS", False),
        ("descriptivas, no verdad humana confirmada. Esta planilla las convierte", False),
        ("en decisiones.", False),
        ("", False),
        ("Qué NO son las columnas de evidencia", True),
        ("Señalan DÓNDE MIRAR. No argumentan que una etiqueta sea correcta.", False),
        ("evidencia_extracto: cita TEXTUAL de un chunk, con chunk_uid y página.", False),
        ("evidencia_por_dominio: si la PALABRA del dominio aparece literalmente", False),
        ("en el documento y dónde. Es una búsqueda de esa palabra, no un léxico", False),
        ("de dominio ni un juicio. «sin coincidencia literal» también informa.", False),
        ("", False),
        ("REGLAS DE LLENADO", True),
        ("", False),
        ("confirmar", True),
        ("    Dejar las cuatro columnas de corrección VACÍAS.", False),
        ("", False),
        ("corregir", True),
        ("    Completar ÚNICAMENTE los campos que cambian. Los que quedan", False),
        ("    vacíos NO se interpretan como borrados.", False),
        ("", False),
        ("dudoso", True),
        ("    Explicar la duda en observaciones.", False),
        ("", False),
        ("excluir_del_corpus", True),
        ("    Explicar OBLIGATORIAMENTE el motivo en observaciones.", False),
        ("", False),
        ("⚠️ Un campo de corrección VACÍO conserva el valor actual del catálogo.", True),
        ("    Vacío significa «no lo toqué», nunca «bórralo».", False),
        ("    Para dominios múltiples, usar el mismo formato del catálogo:", False),
        ("    separados por barra vertical, sin espacios → contable|regulatorio", False),
        ("", False),
        ("EJEMPLO — fila FICTICIA, no está entre los 24 documentos", True),
        ("", False),
        ("__EJEMPLO__", False),
        ("", False),
        ("Se corrigió solo el tipo. emisor y dominios quedan vacíos, así que", False),
        ("conservan lo que ya dice el catálogo.", False),
        ("", False),
        ("Otras hojas", True),
        ("defectos_de_ingesta: caracteres corruptos detectados en el corpus. NO", False),
        ("    se repararon: son un defecto de ingesta, no una pregunta de", False),
        ("    etiquetado, y se registran aparte para tratarlos como tales.", False),
        ("contexto_automatico: lo que decidió el clasificador. Mirarlo ANTES de", False),
        ("    decidir puede anclar la revisión hacia coincidir con la máquina.", False),
        ("", False),
        ("Nada se aplica al catálogo automáticamente. Esta planilla no escribe", False),
        ("en data/catalog/ ni en PostgreSQL.", False),
        ("", False),
        (f"receta: {manifest['receta']}", False),
        (f"huella del catálogo: {manifest['huella_catalogo']}", False),
    ]

    # The fictional example is rendered as a small table, and it lives ONLY
    # here. Putting it among the 24 rows would put a document that does not
    # exist into the corpus under review.
    # [ES] El ejemplo ficticio se dibuja como una tabla chica, y vive SOLO aca.
    # Ponerlo entre las 24 filas metería un documento inexistente en el corpus
    # bajo revision.
    EJEMPLO_ENCABEZADOS = ("document_id", "decision_humana") + CAMPOS_CORREGIBLES + ("observaciones",)
    EJEMPLO_FILA = (
        "DOC-9999 (ficticio)", "corregir", "", "",
        "resolucion_general", "", "El texto se dicta como RG, no como resolución simple.",
    )

    fila_actual = 1
    for texto, negrita in instrucciones:
        if texto == "__EJEMPLO__":
            for j, nombre in enumerate(EJEMPLO_ENCABEZADOS, start=1):
                celda = hoja.cell(row=fila_actual, column=j, value=nombre)
                celda.font = Font(bold=True, size=9)
                celda.fill = PatternFill("solid", fgColor="D9D9D9")
                celda.alignment = Alignment(wrap_text=True, vertical="top")
            fila_actual += 1
            for j, valor in enumerate(EJEMPLO_FILA, start=1):
                celda = hoja.cell(row=fila_actual, column=j, value=valor)
                celda.font = Font(size=9, italic=True)
                celda.fill = PatternFill("solid", fgColor="FFF2CC")
                celda.alignment = Alignment(wrap_text=True, vertical="top")
            fila_actual += 1
            continue
        celda = hoja.cell(row=fila_actual, column=1, value=texto)
        if negrita:
            celda.font = Font(bold=True, size=12 if fila_actual == 1 else 11)
        fila_actual += 1

    hoja.column_dimensions["A"].width = 74
    for j in range(2, len(EJEMPLO_ENCABEZADOS) + 1):
        hoja.column_dimensions[get_column_letter(j)].width = 22

    # --- the review sheet / la hoja de revision ---
    rev = libro.create_sheet("revision")
    rev.append(list(COLUMNAS))

    encabezado = Font(bold=True, color="FFFFFF")
    fondo = PatternFill("solid", fgColor="2F5597")
    fondo_completar = PatternFill("solid", fgColor="FFF2CC")

    for j, nombre in enumerate(COLUMNAS, start=1):
        celda = rev.cell(row=1, column=j)
        celda.font = encabezado
        celda.fill = fondo
        celda.alignment = Alignment(vertical="center", wrap_text=True)

    for fila in filas:
        rev.append([fila[c] for c in COLUMNAS])

    anchos = {
        "document_id": 12, "fuente": 30, "titulo_oficial": 42,
        "archivo_referencia": 30, "emisor_id": 11, "emisor_nombre": 26,
        "tipo_documento": 22, "dominios_documentales": 30, "jurisdiccion": 18,
        "fecha_documento": 14, "chunks": 8, "paginas": 10,
        "evidencia_extracto": 70, "evidencia_por_dominio": 46, "url_origen": 26,
        "estado_inclusion": 20, "decision_humana": 20,
        "emisor_id_corregido": 20, "emisor_nombre_corregido": 28,
        "tipo_documento_corregido": 26, "dominios_documentales_corregidos": 34,
        "observaciones": 44,
    }
    for j, nombre in enumerate(COLUMNAS, start=1):
        rev.column_dimensions[get_column_letter(j)].width = anchos.get(nombre, 18)

    for i in range(2, len(filas) + 2):
        rev.row_dimensions[i].height = 96
        for j, nombre in enumerate(COLUMNAS, start=1):
            celda = rev.cell(row=i, column=j)
            celda.alignment = Alignment(vertical="top", wrap_text=True)
            if nombre in COLUMNAS_A_COMPLETAR:
                celda.fill = fondo_completar

    validacion = DataValidation(
        type="list",
        formula1='"' + ",".join(OPCIONES_DECISION) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    rev.add_data_validation(validacion)
    columna_decision = get_column_letter(COLUMNAS.index("decision_humana") + 1)
    validacion.add(f"{columna_decision}2:{columna_decision}{len(filas) + 1}")

    rev.freeze_panes = "C2"
    rev.auto_filter.ref = (
        f"A1:{get_column_letter(len(COLUMNAS))}{len(filas) + 1}"
    )

    # --- automatic context, deliberately apart / contexto automatico, aparte ---
    ctx = libro.create_sheet("contexto_automatico")
    ctx.append(["⚠️ NO es verdad de referencia. Es lo que decidió el clasificador."])
    ctx.append(["Mirarlo ANTES de decidir puede anclar la revisión hacia coincidir"])
    ctx.append(["con la máquina. Sugerencia: completá 'revision' primero."])
    ctx.append([])
    encabezados_ctx = list(contexto[0].keys()) if contexto else []
    ctx.append(encabezados_ctx)
    for fila in contexto:
        ctx.append([fila[c] for c in encabezados_ctx])
    for j, nombre in enumerate(encabezados_ctx, start=1):
        ctx.column_dimensions[get_column_letter(j)].width = 34 if j > 1 else 14
    for i in (1, 2, 3):
        ctx.cell(row=i, column=1).font = Font(bold=(i == 1), color="B00020")

    # --- ingestion defects, registered and NOT repaired ---
    # [ES] Defectos de ingesta, registrados y NO reparados.
    dfx = libro.create_sheet("defectos_de_ingesta")
    dfx.append(["Caracteres de reemplazo (U+FFFD) encontrados en el corpus."])
    dfx.append(["NO se repararon. Son un defecto de INGESTA, no una pregunta de"])
    dfx.append(["etiquetado: un byte que no se pudo decodificar al ingerir. Arreglarlo"])
    dfx.append(["acá taparía el problema y la próxima reingesta lo traería igual."])
    dfx.append([])
    if defectos:
        encabezados_dfx = list(defectos[0].keys())
        dfx.append(encabezados_dfx)
        for fila in defectos:
            dfx.append([fila[c] for c in encabezados_dfx])
        for j, nombre in enumerate(encabezados_dfx, start=1):
            dfx.column_dimensions[get_column_letter(j)].width = (
                90 if nombre == "contexto" else 20
            )
    else:
        dfx.append(["(ninguno detectado)"])
    for i in range(1, 5):
        dfx.cell(row=i, column=1).font = Font(bold=(i == 1), color="B00020")

    ruta.parent.mkdir(parents=True, exist_ok=True)
    libro.save(ruta)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument(
        "--salida", type=Path, default=EXPERIMENTS_DIR / "revision_catalogo_24"
    )
    args = parser.parse_args()

    artifacts = leer_inventario()
    catalogo = leer_catalogo()
    chunks_por_doc = leer_chunks(artifacts)

    filas = armar_filas(catalogo, chunks_por_doc)
    contexto = contexto_automatico(catalogo, chunks_por_doc)
    defectos = defectos_de_ingesta(chunks_por_doc)

    csv_ruta = args.salida / "revision_catalogo_24_v2.csv"
    xlsx_ruta = args.salida / "revision_catalogo_24_v2.xlsx"

    manifest = {
        "receta": RECETA_VERSION,
        "documentos": len(filas),
        "chunks_considerados": sum(f["chunks"] for f in filas),
        "columnas": list(COLUMNAS),
        "columnas_a_completar": list(COLUMNAS_A_COMPLETAR),
        "campos_corregibles": list(CAMPOS_CORREGIBLES),
        "opciones_decision": list(OPCIONES_DECISION),
        "regla_de_campos_vacios": (
            "un campo de correccion vacio conserva el valor actual del catalogo; "
            "vacio significa 'no lo toque', nunca 'borralo'"
        ),
        "defectos_de_ingesta_detectados": len(defectos),
        "dominios_del_sistema": sorted(SILOS),
        "huella_catalogo": huella_de_archivo(CATALOGO),
        "huella_inventario": huella_de_archivo(INVENTARIO),
        "salvedades": [
            "Este instrumento NO cambia ninguna etiqueta ni escribe en el catalogo.",
            "Las columnas de evidencia senalan donde mirar; no justifican una etiqueta.",
            "`evidencia_por_dominio` es una busqueda literal de la palabra del dominio, "
            "no un lexico de dominio ni un juicio.",
            "La clasificacion automatica (`chunks.silo`) va en una hoja aparte para no "
            "anclar la revision.",
            "Los caracteres de reemplazo (U+FFFD) se REGISTRAN en su propia hoja y NO se "
            "reparan: son defecto de ingesta, no pregunta de etiquetado.",
            "El ejemplo de llenado es ficticio y vive SOLO en la hoja de instrucciones; "
            "no esta entre los 24 documentos.",
            "Solo lectura: no se modifico PostgreSQL, ni la ingesta, ni los embeddings, "
            "ni ningun chunk_uid.",
        ],
    }

    escribir_csv(csv_ruta, filas)
    escribir_xlsx(xlsx_ruta, filas, contexto, defectos, manifest)
    (args.salida / "manifest_v2.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    (args.salida / "defectos_de_ingesta.json").write_text(
        json.dumps(defectos, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )

    sin_coincidencia = [
        f["document_id"] for f in filas
        if "sin coincidencia literal" in f["evidencia_por_dominio"]
    ]

    print(f"documentos           {len(filas)}")
    print(f"defectos de ingesta  {len(defectos)} (registrados, NO reparados)")
    print(f"chunks considerados  {sum(f['chunks'] for f in filas)}")
    print(f"con algun dominio sin coincidencia literal: {len(sin_coincidencia)} "
          f"({', '.join(sin_coincidencia) or '—'})")
    print()
    print(f"xlsx      {xlsx_ruta}")
    print(f"csv       {csv_ruta}")
    print(f"manifest  {args.salida / 'manifest_v2.json'}")
    print(f"defectos  {args.salida / 'defectos_de_ingesta.json'}")


if __name__ == "__main__":
    main()
